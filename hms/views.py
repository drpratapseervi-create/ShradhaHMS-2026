import os
import csv
import logging
import json
from decimal import Decimal
from datetime import date
from django.http import JsonResponse
from openai import OpenAI
from django.conf import settings

from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from .models import PrescriptionTemplate, PrescriptionTemplateItem
from django.http import HttpResponse, JsonResponse
from decimal import Decimal
from django.db.models import Sum
from .models import PartnerDeposit
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.db import transaction
from django.db.models import Q, F, Sum, Count
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .decorators import role_required
# PDF Generation
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from .models import USGReport
from .forms  import USGReportForm
from django.utils import timezone
from .models import (
    Ward, Bed, IPDAdmission, IPDVital, IPDMedication, IPDDischargeMedication, DischargeTemplate, IPDProgressNote, IPDSymptomHistory, IPDTreatmentHistory,
    Patient, Doctor, Department, Appointment, Consultation, Prescription,
    Investigation, InvestigationCategory, InvestigationBill, InvestigationBillItem,
    InvestigationResult, InvestigationParameter, ICDCode, DrugMaster, VillageMaster,
    Symptom, Sign, PastHistory, SurgicalHistory, AdviceOption, DietAdviceOption, MedicalImage, BillItem, PatientService,
    DischargeBill, DischargeBillItem, ProcedureItem, ProcedureBill, ProcedureBillItem,
    IPDAdvance, OTBooking, OTNotes,                        # ← OT models here
    InventoryItem, StockIn, StockOut, Supplier,            # ← Inventory models here
    HospitalDocument,
    HOSPITAL_DOC_TYPES, DOCTOR_DOC_TYPES, STAFF_DOC_TYPES, EQUIPMENT_DOC_TYPES
)

from .models import (
    ConstructionExpense, Vendor, PartnerPayment, ExpenseBudget,
    EXPENSE_HEAD_CHOICES, AREA_CHOICES, PAYMENT_MODE_CHOICES,
    PAID_BY_CHOICES, PAID_FROM_CHOICES, APPROVAL_STATUS_CHOICES,
    APPROVED_BY_CHOICES, WORK_STATUS_CHOICES, YES_NO_PARTIAL_CHOICES,
    INVOICE_TYPE_CHOICES,
)

# Forms
from .forms import (
    PatientForm,
    AppointmentForm,
    ConsultationForm,
    IPDVitalForm,
    IPDAdmissionForm,
)

logger = logging.getLogger(__name__)
# ======================================================
# FONT REGISTRATION (UNICODE SAFE - Rs SUPPORT)
# ======================================================
FONT_PATH = os.path.join(os.path.dirname(__file__), "fonts", "DejaVuSans.ttf")
pdfmetrics.registerFont(TTFont("HospitalFont", FONT_PATH))


# ======================================================
# AMOUNT IN WORDS (INDIAN FORMAT)
# ======================================================
def amount_in_words(number):
    ones = (
        "", "One", "Two", "Three", "Four", "Five", "Six",
        "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve",
        "Thirteen", "Fourteen", "Fifteen", "Sixteen",
        "Seventeen", "Eighteen", "Nineteen",
    )
    tens = (
        "", "", "Twenty", "Thirty", "Forty", "Fifty",
        "Sixty", "Seventy", "Eighty", "Ninety",
    )

    def convert(n):
        if n < 20:
            return ones[n]
        elif n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")
        elif n < 1000:
            return ones[n // 100] + " Hundred" + (
                " " + convert(n % 100) if n % 100 else ""
            )
        elif n < 100000:
            return convert(n // 1000) + " Thousand" + (
                " " + convert(n % 1000) if n % 1000 else ""
            )
        elif n < 10000000:
            return convert(n // 100000) + " Lakh" + (
                " " + convert(n % 100000) if n % 100000 else ""
            )
        else:
            return convert(n // 10000000) + " Crore" + (
                " " + convert(n % 10000000) if n % 10000000 else ""
            )

    return f"{convert(int(number))} Rupees Only"


# ======================================================
# AUTO FLAG HELPER
# ======================================================
def compute_flag(value_str, param):
    """Returns 'CR', 'HIGH', 'LOW', or '' based on parameter limits."""
    if param.result_type != "numeric":
        return ""
    try:
        v = float(value_str)
    except (ValueError, TypeError):
        return ""
    if param.critical_low is not None and v < param.critical_low:
        return "CR"
    if param.critical_high is not None and v > param.critical_high:
        return "CR"
    if param.max_value is not None and v > param.max_value:
        return "HIGH"
    if param.min_value is not None and v < param.min_value:
        return "LOW"
    return ""


# ======================================================
# LAB REPORT BUILDER  (plain clinical print template)
# ======================================================
_LAB_FLAG_LABEL = {"HIGH": "HIGH", "LOW": "LOW", "CR": "CRIT"}


def _lab_num(x):
    """Trim trailing .0 from float range bounds -> '140' not '140.0'."""
    try:
        return f"{float(x):g}"
    except (TypeError, ValueError):
        return str(x)


def _lab_range_str(param, gender):
    if gender == "Female" and param.female_range:
        return param.female_range.strip()
    if gender == "Male" and param.male_range:
        return param.male_range.strip()
    if param.min_value is not None and param.max_value is not None:
        return f"{_lab_num(param.min_value)} - {_lab_num(param.max_value)}"
    if param.min_value is not None:
        return f"> {_lab_num(param.min_value)}"
    if param.max_value is not None:
        return f"< {_lab_num(param.max_value)}"
    return "--"


def _lab_param_code(param, item):
    """LOINC code on the parameter, else the investigation's panel LOINC code, else ''."""
    code = (getattr(param, "loinc_code", "") or "").strip()
    if code:
        return code
    return (getattr(item.investigation, "loinc_panel_code", "") or "").strip()


def _lab_panel_label(param, item):
    grp = (param.group or "").strip()
    if grp:
        return grp.upper()
    try:
        return item.investigation.category.get_dept_code_display().upper()
    except Exception:
        return "INVESTIGATIONS"


# Default reporting/validating doctor for lab reports — this hospital has no
# separate lab-technician field, so all reports are signed off by the lab head.
LAB_REPORTING_DOCTOR = {"name": "Dr. Pratap Senecha, MS, F.MAS",
                        "title": "Reported & Validated By"}


def build_lab_report_context(item, results):
    """
    Structured context for the plain clinical lab-report template.

    Returns a dict with:
      panels      -> [ {"name": "<DEPT>", "rows": [row, ...]}, ... ]
      meta_rows   -> [ (l_label, l_value, r_label, r_value), ... ]  (borderless 2-col)
      remarks     -> single plain-paragraph string
      sig_left / sig_right   -> {"name": ..., "title": ...}  (real names, no placeholder)

    row = {name, code, method, value, flag ('' | HIGH | LOW | CRIT),
           abnormal (bool), range, units}
    method here is the full descriptive sentence (InvestigationParameter.method_description,
    falling back to the short .method label if no description is set).
    """
    patient = item.bill.patient
    gender  = patient.gender or ""

    try:
        ref_doc = item.bill.consultation.appointment.doctor
    except Exception:
        ref_doc = None

    # ---- group rows by panel, preserving encounter order ----
    panel_order, panel_map = [], {}
    for r in results:
        label = _lab_panel_label(r.parameter, item)
        if label not in panel_map:
            panel_map[label] = []
            panel_order.append(label)
        panel_map[label].append({
            "name":     r.parameter.name.strip(),
            "code":     _lab_param_code(r.parameter, item),      # LOINC or panel code or ''
            "method":   ((r.parameter.method_description or "").strip()
                         or (r.parameter.method or "").strip()),
            "value":    str(r.value).strip(),
            "flag":     _LAB_FLAG_LABEL.get(r.flag or "", ""),
            "abnormal": (r.flag or "") in ("HIGH", "LOW", "CR"),
            "range":    _lab_range_str(r.parameter, gender),
            "units":    (r.parameter.unit or "").strip() or "-",
        })
    panels = [{"name": p, "rows": panel_map[p]} for p in panel_order]

    # ---- metadata (borderless two-column) ----
    created = timezone.localtime(item.bill.created_at) if item.bill.created_at else None
    ts = created.strftime("%d-%b-%Y %I:%M %p") if created else "-"

    last_entered = max((r.entered_at for r in results if r.entered_at), default=None)
    reported_ts = (timezone.localtime(last_entered).strftime("%d-%b-%Y %I:%M %p")
                   if last_entered else "-")

    total_params = InvestigationParameter.objects.filter(
        investigation=item.investigation, show_in_report=True).count()
    if not results:
        status = "Pending"
    elif total_params and len(results) >= total_params:
        status = "Final / Validated"
    else:
        status = "Provisional"

    meta_left = [
        ("Patient ID / UHID", patient.uhid or "-"),
        ("Patient Name",      (patient.full_name or "-").title()),
        ("Age / Gender",      f"{patient.age if patient.age is not None else '-'} / {gender or '-'}"),
        ("Referring Dr",      f"Dr. {ref_doc.full_name}" if ref_doc and ref_doc.full_name else "-"),
    ]
    meta_right = [
        ("Bill No",        f"#{item.bill.id}"),
        ("Collected Time", ts),
        ("Reported Time",  reported_ts),
        ("Report Status",  status),
    ]
    # zip into (left_label, left_value, right_label, right_value) rows for the template
    meta_rows = []
    for i in range(max(len(meta_left), len(meta_right))):
        lk, lv = meta_left[i]  if i < len(meta_left)  else ("", "")
        rk, rv = meta_right[i] if i < len(meta_right) else ("", "")
        meta_rows.append((lk, lv, rk, rv))

    # ---- clinical remarks: one plain paragraph ----
    remarks = ("Computer-generated report; values are for clinical reference only. "
               "Please correlate clinically and consult the treating physician for "
               "interpretation.")
    if any(row["abnormal"] for p in panels for row in p["rows"]):
        remarks = ("One or more parameters fall outside the stated reference range and "
                   "have been flagged. ") + remarks

    # ---- signatures ----
    # Left: fixed lab head (no separate technician record in this system).
    sig_left = dict(LAB_REPORTING_DOCTOR)
    # Right: the referring physician from the report data, when there is one.
    if ref_doc and ref_doc.full_name:
        sig_right = {"name": f"Dr. {ref_doc.full_name}",
                     "title": ref_doc.specialization or "Referring Physician"}
    else:
        sig_right = {"name": "-", "title": "Referring Physician"}

    return {
        "panels":    panels,
        "meta_rows": meta_rows,
        "remarks":   remarks,
        "sig_left":  sig_left,
        "sig_right": sig_right,
    }


# ======================================================
# PATIENT
# ======================================================
@login_required
@role_required("reception", "admin", "doctor", "nursing")
def patient_create(request):
    if request.method == "POST":
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            if not patient.date_of_birth:
                age_val = request.POST.get("age")
                if age_val:
                    patient.age_years = int(age_val)
            patient.save()

            # ← Check which button was clicked
            if "save_and_book" in request.POST:
                return redirect(f"/appointments/new/?patient={patient.pk}")
            return redirect("hms:patient_update", pk=patient.pk)
        else:
            print(form.errors)
    else:
        form = PatientForm()

    return render(request, "patients/patient_form.html", {
        "form": form,
        "villages": VillageMaster.objects.all(),
    })
# APPOINTMENT
# ======================================================
@login_required
def appointment_create(request):
    form = AppointmentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        appt = form.save(commit=False)
        appt.status = "Scheduled"
        appt.appointment_type = "New"
        if appt.payment_mode != "FREE" and appt.fee and appt.fee > 0:
            appt.is_paid = True
        appt.save()

        # ── NEW: Check which button was clicked ──
        if 'go_to_consultation' in request.POST:
            return redirect("hms:start_consultation", appointment_id=appt.id)
        else:
            return redirect("hms:print_opd", appointment_id=appt.id)

    return render(request, "appointment_form.html", {"form": form})


# ======================================================
# OPD RECEIPT
# ======================================================
@login_required
def print_opd(request, appointment_id):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "doctor"),
        id=appointment_id,
    )
    return render(request, "opd_receipt_a4.html", {
        "appointment": appointment,
        "printed_on": timezone.now(),
    })

# ======================================================
# CONSULTATION
# ======================================================
@login_required
@role_required("doctor", "admin")
def start_consultation(request, appointment_id):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "doctor"),
        id=appointment_id,
    )
    consultation, _ = Consultation.objects.get_or_create(appointment=appointment)
    prescriptions = Prescription.objects.filter(consultation=consultation)

    if request.method == "POST":
        print("ADVICE:", request.POST.get("advice"))
        print("DIET:", request.POST.get("diet_advice"))
        print("FOLLOWUP:", request.POST.get("follow_up_date"))
        form = ConsultationForm(request.POST, instance=consultation)
        if form.is_valid():
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.appointment = appointment
                obj.last_modified_by = request.user.username

                complaints_manual = request.POST.get("complaints_manual", "").strip()
                exam_manual = request.POST.get("exam_manual", "").strip()
                if complaints_manual:
                    obj.chief_complaints = complaints_manual
                if exam_manual:
                    obj.examination = exam_manual

                # ── Save advice, diet, follow-up ──
                obj.advice          = request.POST.get("advice", "").strip()
                obj.procedures_performed = request.POST.get("procedures_performed", "").strip()
                obj.diet_advice     = request.POST.get("diet_advice", "").strip()
                obj.follow_up_date  = request.POST.get("follow_up_date") or None
                obj.follow_up_notes = request.POST.get("follow_up_notes", "").strip()
                obj.custom_investigations = request.POST.get("custom_investigations", "")

                # Free-text entries under Chief Complaints / Examination Findings —
                # this consultation only, never saved to the Symptom / Sign masters.
                obj.custom_symptoms = request.POST.get("custom_symptoms", "").strip()
                obj.custom_signs    = request.POST.get("custom_signs", "").strip()

                obj.surgery_date    = request.POST.get("surgery_date") or None

                # ── Refusal of Admission / LAMA Consent ──
                obj.lama_declined = bool(request.POST.get("lama_declined"))
                obj.lama_diagnosis = request.POST.get("lama_diagnosis", "").strip()
                obj.lama_plan = request.POST.get("lama_plan", "").strip()
                obj.lama_consent_en = request.POST.get("lama_consent_en", "").strip()
                obj.lama_consent_hi = request.POST.get("lama_consent_hi", "").strip()
                obj.lama_attendant_name = request.POST.get("lama_attendant_name", "").strip()
                obj.lama_attendant_relation = request.POST.get("lama_attendant_relation", "").strip()
                obj.lama_signed_name = request.POST.get("lama_signed_name", "").strip()
                lama_signature_data = request.POST.get("lama_signature_data", "").strip()
                if lama_signature_data:
                    obj.lama_signature_data = lama_signature_data
                if obj.lama_declined and not obj.lama_signed_at:
                    obj.lama_signed_at = timezone.now()

                # ── Quick Lab Values (manual outside/patient-reported entry) ──
                obj.quick_lab_values = {
                    "hb":           request.POST.get("qlv_hb", "").strip(),
                    "tlc":          request.POST.get("qlv_tlc", "").strip(),
                    "platelet":     request.POST.get("qlv_platelet", "").strip(),
                    "rbs":          request.POST.get("qlv_rbs", "").strip(),
                    "creatinine":   request.POST.get("qlv_creatinine", "").strip(),
                    "urea":         request.POST.get("qlv_urea", "").strip(),
                    "sgot":         request.POST.get("qlv_sgot", "").strip(),
                    "sgpt":         request.POST.get("qlv_sgpt", "").strip(),
                    "tsh":          request.POST.get("qlv_tsh", "").strip(),
                    "typhoid":      request.POST.get("qlv_typhoid", "").strip(),
                    "mp_test":      request.POST.get("qlv_mp_test", "").strip(),
                    "esr":          request.POST.get("qlv_esr", "").strip(),
                    "other_label":  request.POST.get("qlv_other_label", "").strip(),
                    "other_value":  request.POST.get("qlv_other_value", "").strip(),
                }

                obj.save()

                obj.symptoms.set(request.POST.getlist("symptoms"))
                obj.signs.set(request.POST.getlist("signs"))
                obj.past_history.set(request.POST.getlist("past_history"))
                obj.surgical_history.set(request.POST.getlist("surgical_history"))

                # ── Save multiple ICD codes ──
                icd_ids = request.POST.getlist("icd_codes[]")
                print("ICD IDS RECEIVED:", icd_ids)
                if icd_ids:
                    obj.icd_codes.set(icd_ids)
                else:
                    obj.icd_codes.clear()

                inv_ids = request.POST.getlist("investigations")
                obj.investigations.set(inv_ids)
                obj.lab_advised = bool(inv_ids)

                # ── Re-apply before second save ──
                obj.advice          = request.POST.get("advice", "").strip()
                obj.diet_advice     = request.POST.get("diet_advice", "").strip()
                obj.follow_up_date  = request.POST.get("follow_up_date") or None
                obj.custom_investigations = request.POST.get("custom_investigations", "")
                obj.follow_up_notes = request.POST.get("follow_up_notes", "").strip()
                obj.save()

                if inv_ids:
                    bill, created = InvestigationBill.objects.get_or_create(
                        consultation=obj,
                        defaults={
                            "patient": appointment.patient,
                            "paid": False,
                            "total_amount": 0,
                        },
                    )
                    if created:
                        total = 0
                        for inv in obj.investigations.all():
                            InvestigationBillItem.objects.create(
                                bill=bill,
                                investigation=inv,
                                price=inv.price,
                                added_by="DOCTOR",
                            )
                            total += inv.price
                        bill.total_amount = total
                        bill.save()

                Prescription.objects.filter(consultation=obj).delete()
                medicines = request.POST.getlist("medicine[]")
                doses     = request.POST.getlist("dose[]")
                freqs     = request.POST.getlist("frequency[]")
                durs      = request.POST.getlist("duration[]")
                instrs    = request.POST.getlist("instructions[]")
                atc_codes = request.POST.getlist("atc_code[]")
                for i, m in enumerate(medicines):
                    if m.strip():
                        Prescription.objects.create(
                            consultation=obj,
                            medicine=m,
                            dose=doses[i] if i < len(doses) else "",
                            frequency=freqs[i] if i < len(freqs) else "",
                            duration=durs[i] if i < len(durs) else "",
                            instructions=instrs[i] if i < len(instrs) else "",
                            atc_code=atc_codes[i] if i < len(atc_codes) else "",
                        )

                messages.success(request, "Consultation saved successfully.")
                return redirect("hms:start_consultation", appointment_id=appointment.id)
    else:
        form = ConsultationForm(instance=consultation)

    previous_consultations = Consultation.objects.filter(
        appointment__patient=appointment.patient
    ).exclude(id=consultation.id).select_related(
        "appointment", "appointment__doctor", "diagnosis_icd"
    ).prefetch_related(
        "symptoms", "signs", "icd_codes", "prescriptions"
    ).order_by("-appointment__date", "-appointment__time")

    previous_visits = [
        {
            "id": c.id,
            "date_str": c.appointment.date.strftime("%b %d, %Y"),
            "doctor": c.appointment.doctor.full_name,
            "pulse": c.pulse,
            "bp": c.bp,
            "spo2": c.spo2,
            "weight": c.weight,
            "quick_lab_values": c.quick_lab_values or {},
            "chief_complaints": c.chief_complaints,
            "symptoms": [s.name for s in c.symptoms.all()],
            "examination": c.examination,
            "signs": [s.name for s in c.signs.all()],
            "diagnosis_text": c.diagnosis_text,
            "icd_codes": [f"{i.code} — {i.description}" for i in c.icd_codes.all()],
            "advice": c.advice,
            "diet_advice": c.diet_advice,
            "follow_up_date": c.follow_up_date.strftime("%b %d, %Y") if c.follow_up_date else "",
            "prescriptions": [
                {
                    "medicine": p.medicine,
                    "dose": p.dose,
                    "frequency": p.frequency,
                    "duration": p.duration,
                    "instructions": p.instructions,
                }
                for p in c.prescriptions.all()
            ],
        }
        for c in previous_consultations
    ]

    return render(request, "opd/consultation.html", {
        "appointment":  appointment,
        "consultation": consultation,
        "form":         form,
        "previous_visits": previous_visits,
        "previous_visits_json": json.dumps(previous_visits).replace("<", "\\u003c"),
        "investigations": Investigation.objects.filter(is_active=True),
        "prescriptions":  prescriptions,
        "symptoms": Symptom.objects.filter(
            department=appointment.department, is_active=True
        ),
        "signs": Sign.objects.filter(
            department=appointment.department, is_active=True
        ),
        "past_histories": PastHistory.objects.filter(is_active=True),
        "surgical_histories": SurgicalHistory.objects.filter(is_active=True),
        "advice_options": AdviceOption.objects.filter(is_active=True),
        "diet_options": DietAdviceOption.objects.filter(is_active=True),
        "medical_images": MedicalImage.objects.filter(
            consultation=consultation
        ).order_by("-created_at"),
        "drug_masters": DrugMaster.objects.filter(is_active=True).order_by("sort_order", "category", "name"),
        "icd_quick_codes": ICDCode.objects.filter(sort_order__lt=9999),
        "icd_categories": ICDCode.CATEGORY_CHOICES,
        "ai_enabled": settings.AI_FEATURES_ENABLED,
    })

# ======================================================
# CONSULTATION PDF
# ======================================================
@login_required
def consultation_pdf(request, appointment_id):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "doctor"),
        id=appointment_id,
    )
    consultation = get_object_or_404(Consultation, appointment=appointment)

    def _lines(txt):
        return [ln.strip() for ln in (txt or "").splitlines() if ln.strip()]

    return render(request, "opd/consultation_pdf.html", {
        "appointment":           appointment,
        "consultation":          consultation,
        "chief_complaints_list": consultation.symptoms.all(),
        "examination_list":      consultation.signs.all(),
        "past_history_list":     consultation.past_history.all(),
        "surgical_history_list": consultation.surgical_history.all(),
        "custom_symptoms":       _lines(consultation.custom_symptoms),
        "custom_signs":          _lines(consultation.custom_signs),
        "investigations":        consultation.investigations.all(),
        "prescriptions":         Prescription.objects.filter(consultation=consultation),
    })


def lama_consent_print(request, appointment_id):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "doctor"),
        id=appointment_id,
    )
    consultation = get_object_or_404(Consultation, appointment=appointment)
    return render(request, "opd/lama_consent_print.html", {
        "appointment":   appointment,
        "consultation":  consultation,
    })


# ======================================================
# LAB BILLING
# ======================================================
@login_required
def lab_billing_direct(request, consultation_id=None):
    consultation = patient = bill = None
    ordered_investigations = []

    bill_id = request.GET.get("bill_id")
    if bill_id:
        bill = get_object_or_404(InvestigationBill, id=bill_id)
        consultation = bill.consultation
        patient = bill.patient
        ordered_investigations = bill.items.values_list("investigation_id", flat=True)
    elif consultation_id:
        consultation = get_object_or_404(Consultation, id=consultation_id)
        patient = consultation.appointment.patient
        ordered_investigations = consultation.investigations.values_list("id", flat=True)

    if request.method == "POST":
        if not patient:
            patient = get_object_or_404(Patient, id=request.POST.get("patient_id"))

        inv_ids      = request.POST.getlist("investigations")
        payment_mode = request.POST.get("payment_mode", "CASH")
        discount     = Decimal(request.POST.get("discount", "0") or "0")

        with transaction.atomic():
            if bill:
                bill.items.all().delete()
                bill.total_amount = 0
                bill.save()
            else:
                bill = InvestigationBill.objects.create(
                    consultation=consultation,
                    patient=patient,
                    paid=False,
                    total_amount=0,
                )

            total = 0
            for inv_id in inv_ids:
                inv = get_object_or_404(Investigation, id=inv_id)
                InvestigationBillItem.objects.create(
                    bill=bill, investigation=inv,
                    price=inv.price, added_by="LAB",
                )
                total += inv.price

            bill.total_amount = total
            bill.discount     = discount
            bill.net_amount   = max(total - discount, Decimal("0"))
            bill.payment_mode = payment_mode
            bill.paid = True
            bill.save()

        return redirect("hms:lab_bill_print", bill_id=bill.id)

    return render(request, "lab/direct_billing.html", {
        "consultation":           consultation,
        "patient":                patient,
        "patients":               Patient.objects.all(),
        "categories":             InvestigationCategory.objects.all(),
        "investigations":         Investigation.objects.filter(is_active=True),
        "ordered_investigations": list(ordered_investigations),
    })


# ======================================================
# PENDING LAB ORDERS
# ======================================================
@login_required
def pending_lab_orders(request):
    q = request.GET.get("q", "").strip()
    pending_bills = (
        InvestigationBill.objects
        .filter(paid=False)
        .select_related("patient", "consultation__appointment")
        .prefetch_related("items__investigation")
        .order_by("-created_at")
    )
    if q:
        filters = Q(patient__full_name__icontains=q) | Q(patient__uhid__icontains=q)
        bill_id_query = q.lstrip("#")
        if bill_id_query.isdigit():
            filters |= Q(id=int(bill_id_query))
        pending_bills = pending_bills.filter(filters)

    paginator = Paginator(pending_bills, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    elided_page_range = list(
        paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)
    )
    return render(request, "lab/pending_lab_orders.html", {
        "pending_bills": page_obj,
        "page_obj": page_obj,
        "elided_page_range": elided_page_range,
        "q": q,
    })


# ======================================================
# MARK LAB BILL AS PAID
# ======================================================
@login_required
def lab_mark_paid(request, bill_id):
    bill = get_object_or_404(InvestigationBill, id=bill_id)
    if request.method == "POST":
        bill.paid = True
        bill.payment_mode = request.POST.get("payment_mode", "CASH")
        bill.save()
        messages.success(request, f"Bill #{bill.id} marked as paid.")
    return redirect("hms:pending_lab_orders")


# ======================================================
# LAB RESULT ENTRY
# ======================================================
@login_required
@role_required("laboratory", "admin")
def lab_result_entry(request, bill_item_id):
    item = get_object_or_404(InvestigationBillItem, id=bill_item_id)

    if not item.bill.paid:
        messages.error(request, "Cannot enter results for unpaid bill.")
        return redirect("hms:pending_lab_orders")

    parameters = InvestigationParameter.objects.filter(
        investigation=item.investigation
    ).order_by("order")

    existing_results = {
        r.parameter_id: r.value
        for r in InvestigationResult.objects.filter(bill_item=item)
    }

    if request.method == "POST":
        InvestigationResult.objects.filter(bill_item=item).delete()
        results_to_create = []
        for p in parameters:
            value = request.POST.get(f"param_{p.id}", "").strip()
            if value:
                results_to_create.append(
                    InvestigationResult(
                        bill_item=item,
                        parameter=p,
                        value=value,
                        entered_by=request.user.username,
                    )
                )
        if results_to_create:
            InvestigationResult.objects.bulk_create(results_to_create)
            messages.success(request, "Lab results saved successfully.")
        return redirect("hms:lab_report_print", bill_item_id=item.id)

    return render(request, "lab/result_entry.html", {
        "item":             item,
        "parameters":       parameters,
        "existing_results": existing_results,
    })


# ======================================================
# LAB REPORTS LIST
# ======================================================
@login_required
def lab_reports(request):
    completed_items = (
        InvestigationBillItem.objects
        .filter(bill__paid=True)
        .select_related("bill__patient", "investigation")
        .prefetch_related("results")
        .order_by("-bill__created_at")
        .distinct()
    )
    return render(request, "lab/lab_reports.html", {"completed_items": completed_items})


# ======================================================
# LAB REPORT PRINT
# ======================================================
@login_required
def lab_report_print(request, bill_item_id):
    item = get_object_or_404(
        InvestigationBillItem.objects.select_related(
            "bill__patient",
            "bill__consultation__appointment__doctor",
            "investigation__category",
        ),
        id=bill_item_id,
    )

    results = list(
        InvestigationResult.objects
        .filter(bill_item=item)
        .select_related("parameter")
        .order_by("parameter__order", "parameter__name")
    )

    for r in results:
        r.flag = compute_flag(r.value, r.parameter)

    ctx = {"item": item, "results": results}
    ctx.update(build_lab_report_context(item, results))
    return render(request, "lab/report_print.html", ctx)


# ======================================================
# LAB BILL PRINT
# ======================================================
@login_required
def lab_bill_print(request, bill_id):
    bill  = get_object_or_404(InvestigationBill, id=bill_id)
    items = InvestigationBillItem.objects.filter(bill=bill)
    net_amount_display = bill.net_amount if bill.net_amount else (bill.total_amount - bill.discount)
    return render(request, "lab/lab_bill_print.html", {
        "bill":                bill,
        "items":               items,
        "amount_in_words":     amount_in_words(bill.total_amount),
        "net_amount_display":  net_amount_display,
    })


# ======================================================
# CSV EXPORT
# ======================================================
@login_required
def export_opd_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="opd_register.csv"'
    writer = csv.writer(response)
    writer.writerow(["ID", "Date", "Patient", "Doctor"])
    for a in Appointment.objects.select_related("patient", "doctor"):
        writer.writerow([
            a.id, a.date,
            a.patient.full_name,
            a.doctor.full_name if a.doctor else "",
        ])
    return response


# ======================================================
# MEDICINE SEARCH (AJAX)
# ======================================================
@login_required
def medicine_search(request):
    q = request.GET.get("q", "")
    by_id = request.GET.get("by_id") == "1"
    medicines = DrugMaster.objects.filter(
        Q(name__icontains=q) | Q(generic_name__icontains=q),
        is_active=True
    )[:20]
    results = []
    for m in medicines:
        results.append({
            "id":           m.id if by_id else m.name,
            "text":         m.name,
            "name":         m.name,
            "generic_name": m.generic_name,
            "strength":     m.strength,
        })
    return JsonResponse(results, safe=False)

# ======================================================
# ICD SEARCH (AJAX)
# ======================================================
@login_required
def icd_search(request):
    q = request.GET.get("q", "").strip()
    icds = ICDCode.objects.filter(
        Q(code__icontains=q) | Q(description__icontains=q)
    )[:20]
    return JsonResponse({
        "results": [{"id": i.id, "text": f"{i.code} - {i.description}"} for i in icds]
    })


# ======================================================
# GET DOCTORS (AJAX)
# ======================================================
@login_required
def get_doctors(request):
    dept_id = request.GET.get("department_id")
    doctors = Doctor.objects.filter(department_id=dept_id).values("id", "full_name")
    return JsonResponse({"doctors": list(doctors)})


# ======================================================
# WHATSAPP WEBHOOK
# ======================================================
@csrf_exempt
def whatsapp_webhook(request):
    logger.info("WhatsApp webhook hit")
    return JsonResponse({"status": "ok"})


# ======================================================
# MEDICAL IMAGE UPLOAD (generic - from patient record)
# ======================================================
@login_required
def upload_medical_image(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    if request.method == "POST":
        image = request.FILES.get("image")
        if image:
            MedicalImage.objects.create(
                patient=patient,
                image_type=request.POST.get("image_type"),
                title=request.POST.get("title"),
                image=image,
                report_text=request.POST.get("report_text"),
                uploaded_by=request.user,
            )
            messages.success(request, "Medical image uploaded successfully.")
        return redirect("hms:dashboard")
    return render(request, "medical/upload_image.html", {"patient": patient})


# ======================================================
# MEDICAL IMAGE UPLOAD - linked to a specific consultation
# ======================================================
@login_required
def upload_medical_image_consultation(request, patient_id, consultation_id):
    patient      = get_object_or_404(Patient, id=patient_id)
    consultation = get_object_or_404(Consultation, id=consultation_id)

    if request.method == "POST":
        image_type  = request.POST.get("image_type", "XRAY")
        title       = request.POST.get("title", "").strip()
        image       = request.FILES.get("image")
        report_text = request.POST.get("report_text", "").strip()

        if image and title:
            MedicalImage.objects.create(
                patient      = patient,
                consultation = consultation,
                image_type   = image_type,
                title        = title,
                image        = image,
                report_text  = report_text,
                uploaded_by  = request.user,
            )
            messages.success(request, f"Image '{title}' uploaded successfully.")
        else:
            messages.error(request, "Title and image file are required.")

    return redirect(
        "hms:start_consultation",
        appointment_id=consultation.appointment.id,
    )


# ======================================================
# DELETE MEDICAL IMAGE
# ======================================================
@login_required
def delete_medical_image(request, image_id):
    img = get_object_or_404(MedicalImage, id=image_id)

    appointment_id = request.POST.get("appointment_id")
    redirect_to    = request.POST.get("redirect_to", "radiology")

    if img.image:
        if os.path.isfile(img.image.path):
            os.remove(img.image.path)

    img.delete()
    messages.success(request, "Image deleted successfully.")

    if redirect_to == "consultation" and appointment_id:
        return redirect("hms:start_consultation", appointment_id=appointment_id)

    return redirect("hms:radiology_upload")


# ======================================================
# RADIOLOGY UPLOAD (Lab Staff - standalone)
# ======================================================
@login_required
@role_required("doctor", "admin", "nursing", "laboratory")
def radiology_upload(request):
    from hms.dicom import upload_dicom

    search_query = request.GET.get("q", "").strip()
    if search_query:
        patients = Patient.objects.filter(
            Q(full_name__icontains=search_query) |
            Q(uhid__icontains=search_query) |
            Q(mobile_no__icontains=search_query)
        ).order_by("-id")[:50]
    else:
        patients = Patient.objects.all().order_by("-id")[:50]

    if request.method == "POST":
        patient_id  = request.POST.get("patient_id")
        image_type  = request.POST.get("image_type")
        title       = request.POST.get("title", "").strip()
        image       = request.FILES.get("image")
        report_text = request.POST.get("report_text", "").strip()

        errors = []
        if not patient_id: errors.append("Please select a patient.")
        if not title:       errors.append("Please enter a title.")
        if not image:       errors.append("Please select an image file.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            patient = get_object_or_404(Patient, id=patient_id)

            # ✅ DICOM file (.dcm) — upload to Orthanc
            if image.name.lower().endswith('.dcm'):
                dicom_bytes = image.read()
                instance_id = upload_dicom(dicom_bytes)
                if instance_id:
                    MedicalImage.objects.create(
                        patient           = patient,
                        image_type        = image_type,
                        title             = title,
                        report_text       = report_text,
                        uploaded_by       = request.user,
                        dicom_instance_id = instance_id,
                        is_dicom          = True,
                    )
                    messages.success(request,
                        f"✅ DICOM uploaded to Orthanc for "
                        f"{patient.full_name} ({patient.uhid}) "
                        f"— ID: {instance_id[:8]}..."
                    )
                else:
                    messages.error(request,
                        "DICOM upload to Orthanc failed. "
                        "Make sure Orthanc is running on port 8042."
                    )

            # Regular image (JPG/PNG/GIF/WEBP)
            else:
                allowed_types = ["image/jpeg", "image/png",
                                 "image/gif",  "image/webp"]
                if image.content_type not in allowed_types:
                    messages.error(request,
                        "Only image files (JPG, PNG, GIF, WEBP) "
                        "or DICOM (.dcm) files are allowed."
                    )
                else:
                    MedicalImage.objects.create(
                        patient     = patient,
                        image_type  = image_type,
                        title       = title,
                        image       = image,
                        report_text = report_text,
                        uploaded_by = request.user,
                        is_dicom    = False,
                    )
                    messages.success(request,
                        f"✅ Image uploaded for "
                        f"{patient.full_name} ({patient.uhid})"
                    )

            return redirect("hms:radiology_upload")

    return render(request, "lab/radiology_upload.html", {
        "patients":       patients,
        "search_query":   search_query,
        "recent_uploads": MedicalImage.objects.select_related("patient")
                          .order_by("-created_at")[:15],
    })
# ======================================================
# IPD DASHBOARD
# ======================================================
@login_required
def ipd_dashboard(request):
    wards = Ward.objects.prefetch_related("bed_set").all()
    admissions = IPDAdmission.objects.filter(
        status="ADMITTED"
    ).select_related("patient", "bed", "doctor")
    return render(request, "ipd/dashboard.html", {
        "wards": wards,
        "admissions": admissions,
    })


# ======================================================
# ADMIT INTO A SPECIFIC BED
# ======================================================
@login_required
def admit_bed(request, bed_id):
    bed = get_object_or_404(Bed, id=bed_id)

    if bed.is_occupied:
        return redirect("hms:ipd_dashboard")

    if request.method == "POST":
        patient_id = request.POST.get("patient")
        doctor_id  = request.POST.get("doctor")

        chief_complaint    = request.POST.get("chief_complaint", "")
        symptoms           = request.POST.get("symptoms", "")
        diagnosis          = request.POST.get("diagnosis", "")
        icd_code           = request.POST.get("icd_code", "")
        attendant_name     = request.POST.get("attendant_name", "")
        attendant_relation = request.POST.get("attendant_relation", "")
        attendant_mobile   = request.POST.get("attendant_mobile", "")

        admission_date = timezone.now()
        admission_date_raw = request.POST.get("admission_date", "").strip()
        if admission_date_raw:
            parsed_date = parse_datetime(admission_date_raw)
            if parsed_date:
                if timezone.is_naive(parsed_date):
                    parsed_date = timezone.make_aware(parsed_date, timezone.get_current_timezone())
                admission_date = parsed_date

        if not patient_id:
            patients = Patient.objects.all().order_by("full_name")
            doctors  = Doctor.objects.select_related("department").order_by("full_name")
            return render(request, "ipd/admit_form.html", {
                "bed": bed,
                "patients": patients,
                "doctors": doctors,
                "error": "Please select a patient before admitting."
            })

        patient = get_object_or_404(Patient, id=int(patient_id))

        doctor = None
        department = None
        if doctor_id:
            doctor = Doctor.objects.select_related("department").filter(id=int(doctor_id)).first()
            if doctor and doctor.department:
                department = doctor.department

        auto_diagnosis = diagnosis
        if not auto_diagnosis:
            last_consultation = Consultation.objects.filter(
                appointment__patient=patient
            ).order_by("-created_at").first()
            if last_consultation:
                auto_diagnosis = last_consultation.diagnosis_text

        last_ipd = IPDAdmission.objects.order_by("-id").first()
        if last_ipd and last_ipd.ipd_no:
            last_num = int(last_ipd.ipd_no.split("-")[-1])
            new_ipd = f"IPD-{last_num+1:05d}"
        else:
            new_ipd = "IPD-00001"

        admission = IPDAdmission.objects.create(
            ipd_no             = new_ipd,
            patient            = patient,
            doctor             = doctor,
            department         = department,
            ward               = bed.ward,
            bed                = bed,
            chief_complaint    = chief_complaint,
            symptoms           = symptoms,
            diagnosis          = auto_diagnosis,
            icd_code           = icd_code,
            attendant_name     = attendant_name,
            attendant_relation = attendant_relation,
            attendant_mobile   = attendant_mobile,
            admission_date     = admission_date,
            status             = "ADMITTED",
        )

        bed.is_occupied = True
        bed.save()

        return redirect("hms:ipd_dashboard")

    patients = Patient.objects.all().order_by("full_name")
    doctors  = Doctor.objects.select_related("department").order_by("full_name")
    return render(request, "ipd/admit_form.html", {
        "bed": bed,
        "patients": patients,
        "doctors": doctors,
    })


# ======================================================
# DISCHARGE PATIENT
# ======================================================
@login_required
def ipd_discharge(request, admission_id):
    admission = get_object_or_404(IPDAdmission, id=admission_id)
    admission.discharge_date = timezone.now()
    admission.status = "DISCHARGED"
    admission.save()
    bed = admission.bed
    bed.is_occupied = False
    bed.save()
    return redirect("hms:ipd_dashboard")


# ======================================================
# ADMIT PATIENT (form-based)
# ======================================================
@login_required
@role_required("doctor", "admin", "nursing")
def admit_patient(request):

    if request.method == "POST":
        form = IPDAdmissionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("hms:ipd_dashboard")
        selected_patient_id = request.POST.get("patient")
    else:
        form = IPDAdmissionForm()
        selected_patient_id = request.GET.get("patient")

    try:
        selected_patient_id = int(selected_patient_id)
    except (TypeError, ValueError):
        selected_patient_id = None

    patients = Patient.objects.all().order_by("full_name")
    doctors = Doctor.objects.all().order_by("full_name")
    return render(request, "ipd/admit_form.html", {
        "form": form,
        "patients": patients,
        "doctors": doctors,
        "selected_patient_id": selected_patient_id,
    })


# ======================================================
# IPD PATIENT FILE
# ======================================================
@login_required
def ipd_patient_file(request, admission_id):
    admission = get_object_or_404(IPDAdmission, id=admission_id)

    if request.method == "POST":
        form_type = request.POST.get("form_type")
        
        if form_type == "vitals":
            IPDVital.objects.create(
                admission=admission,
                pulse=request.POST.get("pulse") or None,
                bp=request.POST.get("bp", ""),
                temperature=request.POST.get("temperature") or None,
                spo2=request.POST.get("spo2") or None,
                rr=request.POST.get("rr") or None
            )
            
        elif form_type == "symptoms":
            # 1. Grab all selected inputs from the checklist array
            selected_list = request.POST.getlist("symptoms_list")
            
            # 2. Join checkboxes together into a clean text string 
            symptoms_text = ", ".join(selected_list)
            admission.symptoms = symptoms_text
            admission.save()

            # 3. Log a historical snapshot of this save (only if something was selected)
            if symptoms_text:
                IPDSymptomHistory.objects.create(
                    admission=admission,
                    symptoms=symptoms_text,
                )

        elif form_type == "treatment":
            treatment_text = request.POST.get("treatment_plan", "").strip()
            if treatment_text:
                admission.treatment_plan = treatment_text
                IPDTreatmentHistory.objects.create(
                    admission=admission,
                    treatment_plan=treatment_text,
                )

        elif form_type == "medication":
            drug_id = request.POST.get("drug_id", "").strip()
            drug_obj = DrugMaster.objects.filter(id=drug_id, is_active=True).first() if drug_id.isdigit() else None
            medicine_name = drug_obj.name if drug_obj else drug_id

            if medicine_name:
                IPDMedication.objects.create(
                    admission=admission,
                    drug=drug_obj,
                    medicine_name=medicine_name,
                    dose=request.POST.get("dose", ""),
                    route=request.POST.get("route", ""),
                    frequency=request.POST.get("frequency", ""),
                )

        elif form_type == "discharge_medication":
            drug_id = request.POST.get("drug_id", "").strip()
            drug_obj = DrugMaster.objects.filter(id=drug_id, is_active=True).first() if drug_id.isdigit() else None
            medicine_name = drug_obj.name if drug_obj else drug_id

            if medicine_name:
                dose = request.POST.get("dose", "").strip()
                if not dose and drug_obj:
                    dose = drug_obj.default_dose or drug_obj.strength
                frequency = request.POST.get("frequency", "").strip() or (drug_obj.default_frequency if drug_obj else "")
                duration = request.POST.get("duration", "").strip() or (drug_obj.default_duration if drug_obj else "")
                instructions = request.POST.get("instructions", "").strip() or (drug_obj.default_instructions if drug_obj else "")

                IPDDischargeMedication.objects.create(
                    admission=admission,
                    drug=drug_obj,
                    medicine_name=medicine_name,
                    dose=dose,
                    route=request.POST.get("route", ""),
                    frequency=frequency,
                    duration=duration,
                    instructions=instructions,
                )

        elif form_type == "discharge_medication_delete":
            med_id = request.POST.get("discharge_medication_id", "")
            if med_id.isdigit():
                IPDDischargeMedication.objects.filter(id=med_id, admission=admission).delete()

        elif form_type == "discharge_medication_bulk":
            med_ids = [i for i in request.POST.getlist("medication_ids") if i.isdigit()]
            source_meds = IPDMedication.objects.filter(id__in=med_ids, admission=admission)
            for src in source_meds:
                IPDDischargeMedication.objects.create(
                    admission=admission,
                    drug=src.drug,
                    medicine_name=src.medicine_name,
                    dose=src.dose,
                    route=src.route,
                    frequency=src.frequency,
                )

        elif form_type == "investigations":
            inv_ids = [i for i in request.POST.getlist("investigations") if i.isdigit()]
            if inv_ids:
                bill, _ = InvestigationBill.objects.get_or_create(
                    admission=admission,
                    defaults={"patient": admission.patient, "paid": False, "total_amount": 0},
                )
                already_ordered = set(
                    bill.items.values_list("investigation_id", flat=True)
                )
                for inv in Investigation.objects.filter(id__in=inv_ids).exclude(id__in=already_ordered):
                    InvestigationBillItem.objects.create(
                        bill=bill,
                        investigation=inv,
                        price=inv.price,
                        added_by="DOCTOR",
                    )
                bill.total_amount = sum(bill.items.values_list("price", flat=True))
                bill.save()

        elif form_type == "progress":
            subjective = request.POST.get("subjective", "").strip()
            objective  = request.POST.get("objective", "").strip()
            assessment = request.POST.get("assessment", "").strip()
            plan       = request.POST.get("plan", "").strip()
            if subjective or objective or assessment or plan:
                IPDProgressNote.objects.create(
                    admission=admission,
                    doctor=admission.doctor,
                    subjective=subjective,
                    objective=objective,
                    assessment=assessment,
                    plan=plan,
                )

        elif form_type == "discharge":
            admission.diagnosis               = request.POST.get("diagnosis", "").strip()
            admission.chief_complaint         = request.POST.get("chief_complaint", "").strip()
            admission.general_examination     = request.POST.get("general_examination", "").strip()
            admission.local_examination       = request.POST.get("local_examination", "").strip()
            admission.inv_hb                  = request.POST.get("inv_hb", "").strip()
            admission.inv_tlc                 = request.POST.get("inv_tlc", "").strip()
            admission.inv_platelet_count      = request.POST.get("inv_platelet_count", "").strip()
            admission.inv_rbs                 = request.POST.get("inv_rbs", "").strip()
            admission.inv_hiv                 = request.POST.get("inv_hiv", "").strip()
            admission.inv_hbsag               = request.POST.get("inv_hbsag", "").strip()
            admission.inv_usg                 = request.POST.get("inv_usg", "").strip()
            admission.procedure_done          = request.POST.get("procedure_done", "").strip()
            admission.course_in_hospital      = request.POST.get("course_in_hospital", "").strip()
            admission.condition_at_discharge  = request.POST.get("condition_at_discharge", "").strip()
            admission.treatment_on_discharge  = request.POST.get("treatment_on_discharge", "").strip()
            admission.discharge_advice        = request.POST.get("discharge_advice", "").strip()
            admission.follow_up_date          = request.POST.get("follow_up_date") or None
            admission.follow_up_instructions  = request.POST.get("follow_up_instructions", "").strip()
            admission.discharge_instructions  = request.POST.get("discharge_instructions", "").strip()
            if not admission.discharge_date:
                admission.discharge_date = timezone.now()

        admission.save()
        discharge_med_actions = ("discharge_medication", "discharge_medication_delete", "discharge_medication_bulk")
        tab = "discharge" if form_type in discharge_med_actions else form_type
        anchor = "#discharge-medications-section" if form_type in discharge_med_actions else ""
        return redirect(f"/ipd/patient/{admission.id}/?tab={tab}{anchor}")

    vitals     = IPDVital.objects.filter(admission=admission).order_by("-recorded_at")
    medications = IPDMedication.objects.filter(admission=admission)
    discharge_medications = IPDDischargeMedication.objects.filter(admission=admission).order_by("-created_at")
    drug_masters = DrugMaster.objects.filter(is_active=True).order_by("sort_order", "category", "name")
    discharge_templates = list(
        DischargeTemplate.objects.filter(is_active=True).values(
            "id", "procedure_name", "gender",
            "diagnosis", "chief_complaints", "general_examination", "local_examination",
            "operation_notes", "course_in_hospital", "treatment_on_discharge",
            "advice", "follow_up", "instructions",
        )
    )
    symptom_history = IPDSymptomHistory.objects.filter(admission=admission).order_by("-recorded_at")
    treatment_history = IPDTreatmentHistory.objects.filter(admission=admission).order_by("-recorded_at")

    ordered_investigations = (
        InvestigationBillItem.objects
        .filter(bill__admission=admission)
        .select_related("investigation__category", "bill")
        .prefetch_related("results")
        .order_by("-id")
    )
    ordered_investigation_ids = {item.investigation_id for item in ordered_investigations}
    available_investigations = Investigation.objects.filter(is_active=True).select_related("category")

    # Optional helper: converts text back to a list so boxes stay checked on refresh
    saved_symptoms_list = [s.strip() for s in admission.symptoms.split(",")] if admission.symptoms else []

    return render(request, "ipd/patient_file.html", {
        "admission":                admission,
        "vitals":                   vitals,
        "medications":              medications,
        "discharge_medications":    discharge_medications,
        "drug_masters":             drug_masters,
        "discharge_templates":      discharge_templates,
        "saved_symptoms_list":      saved_symptoms_list,
        "symptom_history":          symptom_history,
        "treatment_history":        treatment_history,
        "available_investigations": available_investigations,
        "ordered_investigations":   ordered_investigations,
        "ordered_investigation_ids": ordered_investigation_ids,
    })

# ======================================================
# DISCHARGE PDF
# ======================================================
@login_required
def discharge_pdf(request, admission_id):
    admission   = get_object_or_404(IPDAdmission, id=admission_id)
    medications = IPDDischargeMedication.objects.filter(admission=admission).order_by("id")
    return render(request, "ipd/discharge_pdf.html", {
        "admission":   admission,
        "patient":     admission.patient,
        "doctor":      admission.doctor,
        "ward":        admission.ward,
        "bed":         admission.bed,
        "medications": medications,
    })


# ======================================================
# PROGRESS NOTES PDF
# ======================================================
@login_required
def progress_notes_pdf(request, admission_id):
    admission = get_object_or_404(IPDAdmission, id=admission_id)
    notes = IPDProgressNote.objects.filter(
        admission=admission
    ).order_by("-created_at")
    return render(request, "ipd/progress_notes_pdf.html", {
        "admission": admission,
        "notes":     notes,
    })


# ======================================================
# DISCHARGE BILL
# ======================================================
@login_required
def discharge_bill(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)

    admission = IPDAdmission.objects.filter(
        patient=patient,
        status="ADMITTED"
    ).order_by("-id").first()

    if not admission:
        return render(request, "ipd/discharge_bill.html", {
            "error": "No active admission found for this patient.",
            "patient": patient,
        })

    bill, created = DischargeBill.objects.get_or_create(
        patient=patient,
        defaults={
            "total_amount": 0,
            "discount": 0,
            "advance_paid": 0,
            "final_amount": 0
        }
    )

    days = (date.today() - admission.admission_date.date()).days
    if days <= 0:
        days = 1

    ward_name = admission.bed.ward.name.split("(")[0].strip()

    def sync_daily_charge(search_name, ward_filter=None):
        query = BillItem.objects.filter(name__icontains=search_name)
        if ward_filter:
            query = query.filter(name__icontains=ward_filter)
        item = query.first()
        if not item:
            return
        line = DischargeBillItem.objects.filter(bill=bill, item=item).first()
        if line:
            line.quantity = days
            line.price    = item.price
            line.total    = item.price * days
            line.save()
        else:
            DischargeBillItem.objects.create(
                bill=bill, item=item,
                quantity=days, price=item.price,
                total=item.price * days
            )

    sync_daily_charge("Bed Charge", ward_name)
    sync_daily_charge("Nursing")
    sync_daily_charge("Consultation")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add_item":
            item_id = request.POST.get("item")
            qty     = int(request.POST.get("quantity", 1))
            item    = get_object_or_404(BillItem, id=item_id)
            line    = DischargeBillItem.objects.filter(bill=bill, item=item).first()
            if line:
                line.quantity += qty
            else:
                line = DischargeBillItem.objects.create(
                    bill=bill, item=item,
                    quantity=qty, price=item.price, total=0
                )
            line.total = line.quantity * line.price
            line.save()

        elif action == "update_payment":
            discount      = Decimal(request.POST.get("discount", "0") or "0")
            bill.discount = discount
            bill.save()

        elif action == "add_advance":
            amount       = Decimal(request.POST.get("advance_amount", "0") or "0")
            payment_mode = request.POST.get("advance_payment_mode", "CASH")
            note         = request.POST.get("advance_note", "")
            if amount > 0:
                advance = IPDAdvance.objects.create(
                    patient=patient,
                    amount=amount,
                    payment_mode=payment_mode,
                    note=note
                )
                return redirect("hms:advance_payment_receipt_single", advance_id=advance.id)

        elif action == "mark_paid":
            payment_mode   = request.POST.get("final_payment_mode", "CASH")
            advances       = IPDAdvance.objects.filter(patient=patient)
            total_advance  = sum(a.amount for a in advances)
            bill_items_now = DischargeBillItem.objects.filter(bill=bill)
            gross          = sum(i.total for i in bill_items_now)
            net            = gross - bill.discount - total_advance
            bill.advance_paid  = total_advance
            bill.payment_mode  = payment_mode
            bill.final_amount  = max(net, Decimal("0"))
            bill.is_paid       = True
            bill.paid_at       = timezone.now()
            bill.save()
            return redirect("hms:final_payment_receipt", patient_id=patient.id)

        return redirect("hms:discharge_bill", patient_id=patient.id)

    bill_items    = DischargeBillItem.objects.filter(bill=bill)
    total_amount  = sum(i.total for i in bill_items)
    advances      = IPDAdvance.objects.filter(patient=patient).order_by("date")
    total_advance = sum(a.amount for a in advances)
    net           = total_amount - bill.discount - total_advance

    bill.total_amount  = total_amount
    bill.advance_paid  = total_advance
    bill.final_amount  = net
    bill.save()

    return render(request, "ipd/discharge_bill.html", {
        "patient":       patient,
        "admission":     admission,
        "items":         BillItem.objects.all(),
        "bill_items":    bill_items,
        "total_amount":  total_amount,
        "bill":          bill,
        "net_amount":    max(net, 0),
        "refund_amount": abs(net) if net < 0 else 0,
        "advances":      advances,
        "total_advance": total_advance,
    })


# ======================================================
# DELETE BILL ITEM
# ======================================================
@login_required
def delete_bill_item(request, item_id):
    item       = get_object_or_404(DischargeBillItem, id=item_id)
    patient_id = item.bill.patient.id
    item.delete()
    return redirect("hms:discharge_bill", patient_id=patient_id)


# ======================================================
# DISCHARGE BILL PDF
# ======================================================
@login_required
def discharge_bill_pdf(request, admission_id):
    admission    = get_object_or_404(IPDAdmission, id=admission_id)
    patient      = admission.patient
    bill         = DischargeBill.objects.get(patient=patient)
    bill_items   = DischargeBillItem.objects.filter(bill=bill)
    total_amount = sum(i.total for i in bill_items)
    bill.refresh_from_db()
    advances      = IPDAdvance.objects.filter(patient=patient).order_by("date")
    total_advance = sum(a.amount for a in advances)
    discount      = bill.discount
    net           = total_amount - discount - total_advance
    net_amount    = max(net, 0)
    refund_amount = abs(net) if net < 0 else 0
    return render(request, "ipd/discharge_bill_pdf.html", {
        "patient":       patient,
        "admission":     admission,
        "bill_items":    bill_items,
        "total_amount":  total_amount,
        "bill":          bill,
        "advances":      advances,
        "total_advance": total_advance,
        "discount":      discount,
        "net_amount":    net_amount,
        "refund_amount": refund_amount,
    })


# ======================================================
# ADVANCE PAYMENT RECEIPT
# ======================================================
@login_required
def advance_payment_receipt(request, patient_id):
    patient   = get_object_or_404(Patient, id=patient_id)
    admission = IPDAdmission.objects.filter(
        patient=patient, status="ADMITTED"
    ).order_by("-id").first()
    bill          = get_object_or_404(DischargeBill, patient=patient)
    bill_items    = DischargeBillItem.objects.filter(bill=bill)
    gross         = sum(i.total for i in bill_items)
    advances      = IPDAdvance.objects.filter(patient=patient).order_by("date")
    total_advance = sum(a.amount for a in advances)
    net           = gross - bill.discount - total_advance
    net_amount    = max(net, 0)
    refund_amount = abs(net) if net < 0 else 0
    receipt_no    = f"ADV-{bill.id:05d}"
    return render(request, "ipd/advance_payment_receipt.html", {
        "patient":       patient,
        "admission":     admission,
        "bill":          bill,
        "now":           timezone.now(),
        "net_amount":    net_amount,
        "refund_amount": refund_amount,
        "receipt_no":    receipt_no,
        "gross_total":   gross,
        "advances":      advances,
        "total_advance": total_advance,
    })


# ======================================================
# ADVANCE PAYMENT RECEIPT (single)
# ======================================================
@login_required
def advance_payment_receipt_single(request, advance_id):
    advance   = get_object_or_404(IPDAdvance, id=advance_id)
    patient   = advance.patient
    admission = IPDAdmission.objects.filter(
        patient=patient, status="ADMITTED"
    ).order_by("-id").first()
    bill          = DischargeBill.objects.filter(patient=patient).first()
    bill_items    = DischargeBillItem.objects.filter(bill=bill) if bill else []
    gross         = sum(i.total for i in bill_items)
    advances      = IPDAdvance.objects.filter(patient=patient).order_by("date")
    total_advance = sum(a.amount for a in advances)
    discount      = bill.discount if bill else 0
    net           = gross - discount - total_advance
    net_amount    = max(net, 0)
    refund_amount = abs(net) if net < 0 else 0
    receipt_no    = f"ADV-{advance.pk:05d}"
    return render(request, "ipd/advance_payment_receipt_single.html", {
        "patient":       patient,
        "admission":     admission,
        "bill":          bill,
        "advance":       advance,
        "receipt_no":    receipt_no,
        "now":           advance.date,
        "net_amount":    net_amount,
        "refund_amount": refund_amount,
        "gross_total":   gross,
        "total_advance": total_advance,
    })


# ======================================================
# FINAL PAYMENT RECEIPT
# ======================================================
@login_required
def final_payment_receipt(request, patient_id):
    patient   = get_object_or_404(Patient, id=patient_id)
    admission = IPDAdmission.objects.filter(patient=patient).order_by("-id").first()
    bill      = get_object_or_404(DischargeBill, patient=patient)
    bill_items    = DischargeBillItem.objects.filter(bill=bill)
    gross         = sum(i.total for i in bill_items)
    advances      = IPDAdvance.objects.filter(patient=patient).order_by("date")
    total_advance = sum(a.amount for a in advances)
    bill.advance_paid = total_advance
    bill.save()
    net           = gross - bill.discount - total_advance
    net_amount    = max(net, Decimal("0"))
    refund_amount = abs(net) if net < 0 else Decimal("0")
    receipt_no    = f"FPR-{bill.id:05d}"
    return render(request, "ipd/final_payment_receipt.html", {
        "patient":        patient,
        "admission":      admission,
        "bill":           bill,
        "bill_items":     bill_items,
        "gross":          gross,
        "net_amount":     net_amount,
        "refund_amount":  refund_amount,
        "receipt_no":     receipt_no,
        "advances_count": advances.count(),
        "now":            bill.paid_at or timezone.now(),
    })


# ======================================================
# PROCEDURE BILLING
# ======================================================
@login_required
def procedure_billing(request):
    patients    = Patient.objects.all().order_by("-id")[:100]
    procedures  = ProcedureItem.objects.filter(is_active=True)
    departments = Department.objects.all()
    doctors     = Doctor.objects.select_related("department")

    if request.method == "POST":
        patient_id    = request.POST.get("patient_id")
        department_id = request.POST.get("department_id") or None
        doctor_id     = request.POST.get("doctor_id") or None
        payment_mode  = request.POST.get("payment_mode", "CASH")
        discount      = request.POST.get("discount", "0") or "0"
        procedure_ids = [pid for pid in request.POST.getlist("procedures") if pid.isdigit()]

        errors = []
        if not patient_id:
            errors.append("Please select a patient.")
        if not procedure_ids:
            errors.append("Please select at least one procedure.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            patient    = get_object_or_404(Patient, id=patient_id)
            department = Department.objects.filter(id=department_id).first() if department_id else None
            doctor     = Doctor.objects.filter(id=doctor_id).first() if doctor_id else None
            discount   = Decimal(discount)
            selected   = ProcedureItem.objects.filter(id__in=procedure_ids)
            total      = sum((p.price for p in selected), Decimal("0"))
            net        = max(total - discount, Decimal("0"))

            bill = ProcedureBill.objects.create(
                patient=patient,
                department=department,
                consultant=doctor,
                payment_mode=payment_mode,
                total_amount=total,
                discount=discount,
                net_amount=net,
                created_by=request.user,
            )

            for proc in selected:
                ProcedureBillItem.objects.create(
                    bill=bill,
                    procedure=proc,
                    price=proc.price,
                )

            messages.success(request, f"Bill generated for {patient.full_name}")
            return redirect("hms:procedure_bill_print", bill_id=bill.id)

    return render(request, "billing/procedure_billing.html", {
        "patients":    patients,
        "procedures":  procedures,
        "departments": departments,
        "doctors":     doctors,
    })


# ======================================================
# PRINT PROCEDURE BILL
# ======================================================
@login_required
def procedure_bill_print(request, bill_id):
    bill = get_object_or_404(
        ProcedureBill.objects.select_related(
            "patient", "consultant", "department"
        ).prefetch_related("items__procedure"),
        id=bill_id
    )
    return render(request, "billing/procedure_bill_print.html", {"bill": bill})

from .models import ProcedureBill

def procedure_pending_list(request):
    bills = ProcedureBill.objects.all().order_by('-id')
    return render(request, "billing/procedure_pending.html", {
        "bills": bills
    })
# ======================================================
# AUTH HELPERS
# ======================================================
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test


def get_user_role(user):
    if user.is_superuser:
        return "superadmin"
    groups = user.groups.values_list("name", flat=True)
    if "Admin" in groups:       return "admin"
    if "Doctor" in groups:      return "doctor"
    if "Nursing Staff" in groups: return "nursing"
    if "Laboratory" in groups:  return "laboratory"
    if "Reception" in groups:   return "reception"
    return "unknown"


def role_required(*group_names):
    def check(user):
        return user.is_superuser or user.groups.filter(name__in=group_names).exists()
    return user_passes_test(check, login_url="/login/")


# ======================================================
# LOGIN / LOGOUT
# ======================================================
def hms_login(request):
    if request.user.is_authenticated:
        return redirect_by_role(request.user)
    error = None
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user     = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect_by_role(user)
        else:
            error = "Invalid username or password."
    return render(request, "hms/login.html", {"error": error})


def redirect_by_role(user):
    role = get_user_role(user)
    redirects = {
        "superadmin": "/admin/",
        "admin":      "/admin/",
        "doctor":     "/dashboard/doctor/",
        "nursing":    "/dashboard/nursing/",
        "laboratory": "/lab/billing/",
        "reception":  "/dashboard/",
    }
    return redirect(redirects.get(role, "/dashboard/"))


def hms_logout(request):
    logout(request)
    return redirect("/login/")


# ======================================================
# ROLE-BASED DASHBOARDS
# ======================================================
@login_required
def dashboard(request):
    from datetime import date as today_date
    today = today_date.today()
    role  = get_user_role(request.user)
    appointments = Appointment.objects.filter(
        date=today
    ).select_related("patient", "doctor", "consultation").order_by("time")
    return render(request, "dashboard.html", {
        "role":               role,
        "patient_count":      Patient.objects.count(),
        "doctor_count":       Doctor.objects.count(),
        "appt_count":         appointments.count(),
        "appointments":       appointments,
        "latest_appointment": appointments.filter(status="Scheduled").first(),
    })


@login_required
@role_required("Doctor")
def doctor_dashboard(request):
    from datetime import date as today_date
    today = today_date.today()
    appointments = Appointment.objects.filter(
        date=today
    ).select_related("patient", "consultation").order_by("time")
    return render(request, "dashboard.html", {
        "appointments":       appointments,
        "appt_count":         appointments.count(),
        "latest_appointment": appointments.filter(status="Scheduled").first(),
    })


@login_required
@role_required("Nursing Staff")
def nursing_dashboard(request):
    admitted = IPDAdmission.objects.filter(
        status="ADMITTED"
    ).select_related("patient", "bed", "ward")
    beds = Bed.objects.all()
    return render(request, "dashboard.html", {
        "admitted_patients": admitted,
        "beds":              beds,
    })


# ======================================================
# OPD REGISTER
# ======================================================
def opd_register(request):
    today = timezone.now().date()
    appointments = Appointment.objects.filter(
        date=today
    ).select_related("patient", "doctor").order_by("time")
    return render(request, "opd_register.html", {"appointments": appointments})


# ======================================================
# DAILY REPORT
# ======================================================
from django.db.models import Sum
from datetime import datetime
from .models import Expense


def daily_report(request):
    today         = timezone.now().date()
    selected_date = request.GET.get('date')

    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            filter_date = today
    else:
        filter_date = today

    # ---------------- LAB COLLECTION ----------------
    # FIX: added bill__paid=True so only collected (paid) bills count
    def inv_by_dept(dept_code):
        return InvestigationBillItem.objects.filter(
            bill__created_at__date=filter_date,
            bill__paid=True,
            investigation__category__dept_code__iexact=dept_code
        ).aggregate(total=Sum('price'))['total'] or 0

    radiology  = inv_by_dept('RADIOLOGY')
    ecg        = inv_by_dept('ECG')
    cardiology = inv_by_dept('CARDIOLOGY')
    histo      = inv_by_dept('HISTOPATHOLOGY')
    biochem    = inv_by_dept('BIOCHEMISTRY')
    hematology = inv_by_dept('HEMATOLOGY')
    microbio   = inv_by_dept('MICROBIOLOGY')
    endoscopy  = inv_by_dept('ENDOSCOPY')

    lab = radiology + ecg + cardiology + histo + biochem + hematology + microbio + endoscopy

    # ---------------- OPD COLLECTION ----------------
    # FIX: is_paid is now correctly set in appointment_create
    opd = Appointment.objects.filter(
        date=filter_date,
        is_paid=True
    ).aggregate(total=Sum('fee'))['total'] or 0

    # ---------------- PROCEDURE COLLECTION ----------------
    procedure = ProcedureBill.objects.filter(
        created_at__date=filter_date
    ).aggregate(total=Sum('net_amount'))['total'] or 0

    # ---------------- IPD ADVANCE ----------------
    advance = IPDAdvance.objects.filter(
        date__date=filter_date
    ).aggregate(total=Sum('amount'))['total'] or 0

    # ---------------- DISCHARGE COLLECTION ----------------
    discharge = DischargeBill.objects.filter(
        paid_at__date=filter_date,
        is_paid=True
    ).aggregate(total=Sum('final_amount'))['total'] or 0

    # ---------------- EXPENSES ----------------
    expenses = Expense.objects.filter(
        date=filter_date
    ).aggregate(total=Sum('amount'))['total'] or 0

    # ---------------- FINAL CALCULATION ----------------
    total_collection = opd + lab + procedure + advance + discharge
    net_collection   = total_collection - expenses

    return render(request, "daily_report.html", {
        "selected_date":    filter_date,
        "opd":              opd,
        "radiology":        radiology,
        "ecg":              ecg,
        "cardiology":       cardiology,
        "histo":            histo,
        "biochem":          biochem,
        "hematology":       hematology,
        "microbio":         microbio,
        "endoscopy":        endoscopy,
        "lab":              lab,
        "procedure":        procedure,
        "advance":          advance,
        "discharge":        discharge,
        "expenses":         expenses,
        "total_collection": total_collection,
        "net_collection":   net_collection,
    })

    from django.shortcuts import get_object_or_404

@login_required
def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    if request.method == "POST":
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect("hms:patient_update", pk=patient.pk)
    else:
        form = PatientForm(instance=patient)

    return render(request, "patients/patient_form.html", {
        "form": form,
        "villages": VillageMaster.objects.all(),
    })

# All imports at the TOP — never inside or between functions
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import OTBooking, Patient  # remove Doctor — use plain text fields instead

@login_required
def ot_dashboard(request):
    today = date.today()
    ot_list = OTBooking.objects.filter(ot_date=today).order_by("ot_time")
    return render(request, "ot/dashboard.html", {"ot_list": ot_list})


from .models import OTBooking, Patient, Doctor  # add Doctor

@login_required
def ot_create(request):
    if request.method == "POST":
        OTBooking.objects.create(
            patient_id=request.POST.get("patient"),
            uhid=request.POST.get("uhid"),
            surgeon=request.POST.get("surgeon"),
            assistant=request.POST.get("assistant"),
            anesthetist=request.POST.get("anesthetist"),
            procedure=request.POST.get("procedure"),
            ot_date=request.POST.get("ot_date"),
            ot_time=request.POST.get("ot_time"),
            ot_room=request.POST.get("ot_room"),
            case_type=request.POST.get("case_type"),
            anesthesia_type=request.POST.get("anesthesia_type"),
        )
        return redirect("hms:ot_dashboard")

    doctors = Doctor.objects.all().order_by("full_name")  # ← now works
    return render(request, "ot/ot_form.html", {"doctors": doctors})


@login_required
def patient_search_api(request):
    q = request.GET.get("q", "")
    if not q:
        return JsonResponse([], safe=False)
    
    from django.db.models import Q
    patients = Patient.objects.filter(
        Q(uhid__icontains=q) | Q(full_name__icontains=q)  # ← full_name not name
    )[:10]
    
    data = [{"id": p.id, "uhid": p.uhid, "name": p.full_name} for p in patients]
    return JsonResponse(data, safe=False)

@login_required
def ot_detail(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    return render(request, "ot/ot_detail.html", {"booking": booking})

# ── OT EDIT ────────────────────────────────────────────
@login_required
def ot_edit(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    doctors = Doctor.objects.all().order_by("full_name")

    if request.method == "POST":
        booking.surgeon        = request.POST.get("surgeon")
        booking.assistant      = request.POST.get("assistant")
        booking.anesthetist    = request.POST.get("anesthetist")
        booking.procedure      = request.POST.get("procedure")
        booking.ot_date        = request.POST.get("ot_date")
        booking.ot_time        = request.POST.get("ot_time")
        booking.ot_room        = request.POST.get("ot_room")
        booking.case_type      = request.POST.get("case_type")
        booking.anesthesia_type = request.POST.get("anesthesia_type")
        booking.save()
        return redirect("hms:ot_detail", id=id)

    return render(request, "ot/ot_edit.html", {"booking": booking, "doctors": doctors})


# ── OT CANCEL ──────────────────────────────────────────
@login_required
def ot_cancel(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    if request.method == "POST":
        booking.status = "Cancelled"
        booking.save()
        return redirect("hms:ot_dashboard")
    return render(request, "ot/ot_confirm_cancel.html", {"booking": booking})


# ── OT STATUS UPDATE ───────────────────────────────────
@login_required
def ot_status_update(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    if request.method == "POST":
        new_status = request.POST.get("status")
        if new_status in ["Scheduled", "Completed", "Cancelled"]:
            booking.status = new_status
            booking.save()
    return redirect("hms:ot_dashboard")


# ── OT NOTES ───────────────────────────────────────────
@login_required
def ot_notes(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    
    # ← use filter().first() instead of get_or_create
    notes = OTNotes.objects.filter(booking=booking).first()

    if request.method == "POST":
        if notes is None:
            notes = OTNotes(booking=booking)
        notes.start_time        = request.POST.get("start_time")
        notes.end_time          = request.POST.get("end_time")
        notes.findings          = request.POST.get("findings")
        notes.procedure_done    = request.POST.get("procedure_done")
        notes.complications     = request.POST.get("complications")
        notes.blood_loss        = request.POST.get("blood_loss")
        notes.post_op_condition = request.POST.get("post_op_condition")
        notes.save()
        booking.status = "Completed"
        booking.save()
        return redirect("hms:ot_detail", id=id)

    return render(request, "ot/ot_notes.html", {"booking": booking, "notes": notes})


# ── OT PRINT ───────────────────────────────────────────
@login_required
def ot_print(request, id):
    booking = get_object_or_404(OTBooking, id=id)
    notes = OTNotes.objects.filter(booking=booking).first()
    return render(request, "ot/ot_print.html", {"booking": booking, "notes": notes})

    from .models import InventoryItem, StockIn, StockOut, Supplier

# ── INVENTORY DASHBOARD ────────────────────────────────
@login_required
def inventory_dashboard(request):
    total_items = InventoryItem.objects.count()
    low_stock   = InventoryItem.objects.filter(current_stock__lte=F('minimum_stock'))  # ← F not models.F
    recent_in   = StockIn.objects.select_related('item').order_by('-created_at')[:5]
    recent_out  = StockOut.objects.select_related('item').order_by('-created_at')[:5]
    return render(request, "inventory/dashboard.html", {
        "total_items": total_items,
        "low_stock": low_stock,
        "recent_in": recent_in,
        "recent_out": recent_out,
    })


# ── INVENTORY REPORT ───────────────────────────────────
@login_required
def inventory_report(request):
    items     = InventoryItem.objects.all().order_by("category", "name")
    low_stock = items.filter(current_stock__lte=F('minimum_stock'))  # ← F not models.F
    return render(request, "inventory/report.html", {
        "items": items,
        "low_stock": low_stock,
    })

# ── INVENTORY ITEMS ────────────────────────────────────
@login_required
def inventory_items(request):
    category = request.GET.get("category", "")
    search   = request.GET.get("q", "")
    items    = InventoryItem.objects.select_related("supplier").order_by("name")
    if category:
        items = items.filter(category=category)
    if search:
        items = items.filter(name__icontains=search)
    return render(request, "inventory/items.html", {"items": items, "category": category, "search": search})


@login_required
def inventory_item_new(request):
    suppliers = Supplier.objects.all().order_by("name")
    if request.method == "POST":
        InventoryItem.objects.create(
            name          = request.POST.get("name"),
            category      = request.POST.get("category"),
            unit          = request.POST.get("unit"),
            current_stock = request.POST.get("current_stock", 0),
            minimum_stock = request.POST.get("minimum_stock", 10),
            supplier_id   = request.POST.get("supplier") or None,
        )
        return redirect("hms:inventory_items")
    return render(request, "inventory/item_form.html", {"suppliers": suppliers})


@login_required
def inventory_item_detail(request, id):
    item     = get_object_or_404(InventoryItem, id=id)
    stock_ins  = item.stock_ins.order_by("-date")[:20]
    stock_outs = item.stock_outs.order_by("-date")[:20]
    return render(request, "inventory/item_detail.html", {
        "item": item,
        "stock_ins": stock_ins,
        "stock_outs": stock_outs,
    })


# ── STOCK IN ───────────────────────────────────────────
@login_required
def stock_in_create(request):
    items     = InventoryItem.objects.all().order_by("name")
    suppliers = Supplier.objects.all().order_by("name")
    if request.method == "POST":
        StockIn.objects.create(
            item_id        = request.POST.get("item"),
            supplier_id    = request.POST.get("supplier") or None,
            quantity       = request.POST.get("quantity"),
            batch_no       = request.POST.get("batch_no"),
            expiry_date    = request.POST.get("expiry_date") or None,
            purchase_price = request.POST.get("purchase_price", 0),
            date           = request.POST.get("date"),
            notes          = request.POST.get("notes"),
            created_by     = request.user,
        )
        return redirect("hms:inventory_dashboard")
    return render(request, "inventory/stock_in_form.html", {"items": items, "suppliers": suppliers})


# ── STOCK OUT ──────────────────────────────────────────
@login_required
def stock_out_create(request):
    items = InventoryItem.objects.filter(current_stock__gt=0).order_by("name")
    if request.method == "POST":
        StockOut.objects.create(
            item_id          = request.POST.get("item"),
            quantity         = request.POST.get("quantity"),
            issued_to        = request.POST.get("issued_to"),
            issued_to_detail = request.POST.get("issued_to_detail"),
            date             = request.POST.get("date"),
            notes            = request.POST.get("notes"),
            created_by       = request.user,
        )
        return redirect("hms:inventory_dashboard")
    return render(request, "inventory/stock_out_form.html", {"items": items})


# ── SUPPLIERS ──────────────────────────────────────────
@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by("name")
    return render(request, "inventory/suppliers.html", {"suppliers": suppliers})


@login_required
def supplier_new(request):
    if request.method == "POST":
        Supplier.objects.create(
            name    = request.POST.get("name"),
            contact = request.POST.get("contact"),
            email   = request.POST.get("email"),
            address = request.POST.get("address"),
        )
        return redirect("hms:supplier_list")
    return render(request, "inventory/supplier_form.html")

# =====================================================================
# DOCUMENT MANAGEMENT VIEWS — Add to your existing views.py
# =====================================================================
# Step 1: Add HospitalDocument to your models import at top of views.py:
#   from .models import (..., HospitalDocument)
# Step 2: Paste these views at the END of your views.py
# =====================================================================

from .models import (
    HospitalDocument,
    HOSPITAL_DOC_TYPES, DOCTOR_DOC_TYPES, STAFF_DOC_TYPES, EQUIPMENT_DOC_TYPES,
)


# ── DOCUMENT DASHBOARD ────────────────────────────────────────────────────────
@login_required
def document_dashboard(request):
    today = timezone.now().date()
    from datetime import timedelta

    all_docs = HospitalDocument.objects.all()

    expired       = [d for d in all_docs if d.expiry_status == 'expired']
    expiring_soon = [d for d in all_docs if d.expiry_status == 'expiring_soon']

    return render(request, 'documents/dashboard.html', {
        'total_docs':      all_docs.count(),
        'hospital_count':  all_docs.filter(category='hospital').count(),
        'doctor_count':    all_docs.filter(category='doctor').count(),
        'staff_count':     all_docs.filter(category='staff').count(),
        'equipment_count': all_docs.filter(category='equipment').count(),
        'expired_count':   len(expired),
        'expiring_count':  len(expiring_soon),
        'expired_docs':    expired[:5],
        'expiring_docs':   expiring_soon[:5],
        'recent_docs':     all_docs.order_by('-created_at')[:8],
        'doc_categories':  [
            ('hospital',  'Hospital'),
            ('doctor',    'Doctor'),
            ('staff',     'Staff'),
            ('equipment', 'Equipment AMC / Certificate'),
        ],
    })
# ── DOCUMENT LIST (filtered by category) ─────────────────────────────────────
@login_required
def document_list(request):
    category = request.GET.get('category', '')
    search   = request.GET.get('q', '').strip()
    status   = request.GET.get('status', '')

    docs = HospitalDocument.objects.all().order_by('category', 'expiry_date')

    if category:
        docs = docs.filter(category=category)
    if search:
        docs = docs.filter(
            Q(title__icontains=search) |
            Q(person_name__icontains=search) |
            Q(equipment_name__icontains=search) |
            Q(issued_by__icontains=search)
        )

    # Filter by expiry status (done in Python since it's a property)
    if status:
        docs = [d for d in docs if d.expiry_status == status]

    return render(request, 'documents/document_list.html', {
        'docs':              docs,
        'category':          category,
        'search':            search,
        'status':            status,
        'hospital_doc_types':  HOSPITAL_DOC_TYPES,
        'doctor_doc_types':    DOCTOR_DOC_TYPES,
        'staff_doc_types':     STAFF_DOC_TYPES,
        'equipment_doc_types': EQUIPMENT_DOC_TYPES,
    })


# ── ADD DOCUMENT ──────────────────────────────────────────────────────────────
@login_required
def document_add(request):
    if request.method == 'POST':
        category       = request.POST.get('category')
        doc_type       = request.POST.get('doc_type')
        title          = request.POST.get('title', '').strip()
        person_name    = request.POST.get('person_name', '').strip()
        equipment_name = request.POST.get('equipment_name', '').strip()
        issued_by      = request.POST.get('issued_by', '').strip()
        issue_date     = request.POST.get('issue_date') or None
        expiry_date    = request.POST.get('expiry_date') or None
        notes          = request.POST.get('notes', '').strip()
        doc_file       = request.FILES.get('document_file')

        if not title or not category or not doc_type:
            messages.error(request, 'Category, Document Type and Title are required.')
        else:
            HospitalDocument.objects.create(
                category       = category,
                doc_type       = doc_type,
                title          = title,
                person_name    = person_name or None,
                equipment_name = equipment_name or None,
                issued_by      = issued_by or None,
                issue_date     = issue_date,
                expiry_date    = expiry_date,
                notes          = notes or None,
                document_file  = doc_file,
            )
            messages.success(request, f'Document "{title}" added successfully.')
            return redirect('hms:document_list')

    return render(request, 'documents/document_form.html', {
        'action':              'Add',
        'hospital_doc_types':  HOSPITAL_DOC_TYPES,
        'doctor_doc_types':    DOCTOR_DOC_TYPES,
        'staff_doc_types':     STAFF_DOC_TYPES,
        'equipment_doc_types': EQUIPMENT_DOC_TYPES,
        'category_pre':        request.GET.get('category', ''),
    })


# ── EDIT DOCUMENT ─────────────────────────────────────────────────────────────
@login_required
def document_edit(request, doc_id):
    doc = get_object_or_404(HospitalDocument, id=doc_id)

    if request.method == 'POST':
        doc.category       = request.POST.get('category')
        doc.doc_type       = request.POST.get('doc_type')
        doc.title          = request.POST.get('title', '').strip()
        doc.person_name    = request.POST.get('person_name', '').strip() or None
        doc.equipment_name = request.POST.get('equipment_name', '').strip() or None
        doc.issued_by      = request.POST.get('issued_by', '').strip() or None
        doc.issue_date     = request.POST.get('issue_date') or None
        doc.expiry_date    = request.POST.get('expiry_date') or None
        doc.notes          = request.POST.get('notes', '').strip() or None

        new_file = request.FILES.get('document_file')
        if new_file:
            doc.document_file = new_file

        doc.save()
        messages.success(request, f'Document "{doc.title}" updated successfully.')
        return redirect('hms:document_list')

    return render(request, 'documents/document_form.html', {
        'action':              'Edit',
        'doc':                 doc,
        'hospital_doc_types':  HOSPITAL_DOC_TYPES,
        'doctor_doc_types':    DOCTOR_DOC_TYPES,
        'staff_doc_types':     STAFF_DOC_TYPES,
        'equipment_doc_types': EQUIPMENT_DOC_TYPES,
    })


# ── DELETE DOCUMENT ───────────────────────────────────────────────────────────
@login_required
def document_delete(request, doc_id):
    doc = get_object_or_404(HospitalDocument, id=doc_id)
    if request.method == 'POST':
        title = doc.title
        if doc.document_file:
            try:
                if os.path.isfile(doc.document_file.path):
                    os.remove(doc.document_file.path)
            except Exception:
                pass
        doc.delete()
        messages.success(request, f'Document "{title}" deleted.')
    return redirect('hms:document_list')


# ── AJAX: Get doc types for a category ────────────────────────────────────────
@login_required
def document_types_ajax(request):
    category = request.GET.get('category', '')
    type_map = {
        'hospital':  HOSPITAL_DOC_TYPES,
        'doctor':    DOCTOR_DOC_TYPES,
        'staff':     STAFF_DOC_TYPES,
        'equipment': EQUIPMENT_DOC_TYPES,
    }
    types = type_map.get(category, [])
    return JsonResponse({'types': [{'value': v, 'label': l} for v, l in types]})


client = OpenAI(api_key=settings.OPENAI_API_KEY)

def ai_full_opd(request):
    if request.method == "POST":
        if not settings.AI_FEATURES_ENABLED:
            return JsonResponse({"error": "AI features are not configured on this system."})

        complaints = request.POST.getlist("complaints[]")
        exam = request.POST.getlist("exam[]")
        vitals = request.POST.get("vitals")
        investigations = request.POST.get("investigation_results")

        prompt = f"""
        You are assisting a General Surgeon in OPD in India.

        Symptoms: {', '.join(complaints)}
        Examination: {', '.join(exam)}
        Vitals: {vitals}
        Investigation Results: {investigations}

        Give structured output:

        Probable Diagnosis:
        Required Investigations:
        Final Diagnosis:
        Treatment Plan:
        Prescription:
        Advice:
        Red Flag:

        Keep short and practical.
        """

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        result_text = response.choices[0].message.content or ""

        return JsonResponse({"result": result_text})

        from django.shortcuts import render
from django.http import HttpResponse
import csv
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib import colors

from .models import ProcedureBill, Consultation
# add your models if present
# from .models import LabBill, DischargeBill, AdvancePayment


# ================= DEPARTMENT DISPLAY =================
def get_dept_display(dept_code):
    mapping = {
        "RADIOLOGY": "Radiology",
        "HISTOPATHOLOGY": "Histopathology",
        "BIOCHEMISTRY": "Biochemistry",
        "HEMATOLOGY": "Hematology",
        "MICROBIOLOGY": "Microbiology",
        "CARDIOLOGY": "Cardiology",
        "ECG": "ECG",
        "ENDOSCOPY": "Endoscopy",
        "OTHER": "Other",
    }
    return mapping.get(dept_code, "Other")


# ================= MASTER DATA FUNCTION =================
def get_all_payment_data(request):
    
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    doctor = request.GET.get("doctor")
    dept = request.GET.get("department")

    data = []

    # ================= PROCEDURE =================
    procedures = ProcedureBill.objects.select_related(
        "patient", "consultant", "department"
    )

    if from_date and to_date:
        procedures = procedures.filter(created_at__date__range=[from_date, to_date])

    if doctor:
        procedures = procedures.filter(consultant_id=doctor)

    if dept:
        procedures = procedures.filter(department_id=dept)

    for p in procedures:
        local_p = timezone.localtime(p.created_at)
        data.append({
            "receipt_no": f"PB-{p.id}",
            "date": local_p,
            "time": local_p.strftime("%I:%M %p"),
            "patient": p.patient.full_name,
            "type": "Procedure",
            "doctor": p.consultant.full_name if p.consultant else "",
            "department": p.department.name if p.department else "",
            "amount": float(p.net_amount)
        })

    # ================= OPD =================
    consultations = Consultation.objects.select_related(
        "appointment__patient", "appointment__doctor"
    )

    if from_date and to_date:
        consultations = consultations.filter(created_at__date__range=[from_date, to_date])

    for c in consultations:
        local_created = timezone.localtime(c.created_at)
        data.append({
            "receipt_no": f"OPD-{c.id}",
            "date": local_created,
            "time": local_created.strftime("%I:%M %p"),
            "patient": c.appointment.patient.full_name if c.appointment and c.appointment.patient else "",
            "type": "OPD",
            "doctor": c.appointment.doctor.full_name if c.appointment and c.appointment.doctor else "",
            "department": "OPD",
            "amount": 100
        })

    # ================= LAB (InvestigationBill) =================
    try:
        lab_bills = InvestigationBill.objects.prefetch_related(
            "items__investigation__category"
        ).select_related("patient", "consultation__appointment__doctor")

        if from_date and to_date:
            lab_bills = lab_bills.filter(created_at__date__range=[from_date, to_date])

        for l in lab_bills:
            dept_set = set()
            for item in l.items.all():
                if item.investigation and item.investigation.category:
                    dept_set.add(item.investigation.category.dept_code)
            dept_names = [get_dept_display(code) for code in dept_set]
            dept_display = ", ".join(dept_names) if dept_names else "Laboratory"

            amount = float(l.net_amount) if l.net_amount else float(l.total_amount)
            local_l = timezone.localtime(l.created_at)

            lab_doctor = ""
            if l.consultation and l.consultation.appointment and l.consultation.appointment.doctor:
                lab_doctor = l.consultation.appointment.doctor.full_name
            elif l.patient:
                last_appt = Appointment.objects.filter(patient=l.patient).order_by("-id").first()
                if last_appt and last_appt.doctor:
                    lab_doctor = last_appt.doctor.full_name

            data.append({
                "receipt_no": f"LAB-{l.id}",
                "date": local_l,
                "time": local_l.strftime("%I:%M %p"),
                "patient": l.patient.full_name if l.patient else "",
                "type": "Lab",
                "doctor": lab_doctor,
                "department": dept_display,
                "amount": amount
            })
    except Exception as e:
        logger.error(f"Lab billing fetch error: {e}")

    # ================= IPD ADVANCE =================
    try:
        advances = IPDAdvance.objects.select_related("patient")
        if from_date and to_date:
            advances = advances.filter(date__date__range=[from_date, to_date])
        for a in advances:
            local_date = timezone.localtime(a.date)
            adm = IPDAdmission.objects.filter(patient=a.patient).order_by("-id").first()
            adv_doctor = adm.doctor.full_name if adm and adm.doctor else ""
            adv_dept   = adm.department.name  if adm and adm.department else "IPD"
            data.append({
                "receipt_no": f"ADV-{a.pk:05d}",
                "date": local_date,
                "time": local_date.strftime("%I:%M %p"),
                "patient": a.patient.full_name if a.patient else "",
                "type": "Advance",
                "doctor": adv_doctor,
                "department": adv_dept,
                "amount": float(a.amount)
            })
    except Exception as e:
        logger.error(f"IPD advance fetch error: {e}")

    # ================= DISCHARGE FINAL PAYMENT (FPR) =================
    try:
        discharge_bills = DischargeBill.objects.select_related("patient").filter(is_paid=True)
        if from_date and to_date:
            discharge_bills = discharge_bills.filter(paid_at__date__range=[from_date, to_date])
        for d in discharge_bills:
            paid_at = d.paid_at or d.created_at
            local_paid = timezone.localtime(paid_at)
            adm = IPDAdmission.objects.filter(patient=d.patient).order_by("-id").first()
            dis_doctor = adm.doctor.full_name if adm and adm.doctor else ""
            dis_dept   = adm.department.name  if adm and adm.department else "IPD"
            data.append({
                "receipt_no": f"FPR-{d.id:05d}",
                "date": local_paid,
                "time": local_paid.strftime("%I:%M %p"),
                "patient": d.patient.full_name if d.patient else "",
                "type": "Discharge",
                "doctor": dis_doctor,
                "department": dis_dept,
                "amount": float(d.final_amount)
            })
    except Exception as e:
        logger.error(f"Discharge final payment fetch error: {e}")

    # SORT
    data = sorted(data, key=lambda x: x["date"], reverse=True)

    return data


# ================= MAIN REPORT PAGE =================
@login_required
@role_required("admin")
def all_reports(request):
    data = get_all_payment_data(request)

    type_filter = request.GET.get("type_filter", "")
    if type_filter:
        data = [r for r in data if r["type"].lower() == type_filter.lower()]

    dept_filter = request.GET.get("dept_filter", "")
    if dept_filter:
        df = dept_filter.lower()
        item_level_filters = {
            "xray":     ["x-ray", "x ray", "xray", "chest x", "spine", "kub", "skull"],
            "usg":      ["usg", "ultrasound", "sonography", "usg abdomen", "usg pelvis"],
            "serology": ["hiv", "hbsag", "hcv", "vdrl", "widal", "aso", "crp", "typhoid", "tridot"],
        }
        if df in item_level_filters:
            keywords = item_level_filters[df]
            matching_ids = set()
            for row in data:
                if row["type"] == "Lab":
                    try:
                        bill_id = int(row["receipt_no"].replace("LAB-", ""))
                        items = InvestigationBillItem.objects.filter(
                            bill_id=bill_id
                        ).select_related("investigation")
                        for item in items:
                            name = item.investigation.name.lower() if item.investigation else ""
                            if any(k in name for k in keywords):
                                matching_ids.add(row["receipt_no"])
                                break
                    except Exception:
                        pass
            data = [r for r in data if r["receipt_no"] in matching_ids]
        else:
            data = [r for r in data if df in r["department"].lower()]

    doctor_filter = request.GET.get("doctor_filter", "")
    if doctor_filter:
        data = [r for r in data if doctor_filter.lower() in r["doctor"].lower()]

    total   = sum(x["amount"] for x in data)
    doctors = Doctor.objects.order_by("full_name")

    return render(request, "reports/all_reports.html", {
        "data":          data,
        "total":         total,
        "doctors":       doctors,
        "type_filter":   type_filter,
        "dept_filter":   dept_filter,
        "doctor_filter": doctor_filter,
        "from_date":     request.GET.get("from_date", ""),
        "to_date":       request.GET.get("to_date", ""),
    })


# ================= CSV =================
@login_required
@role_required("admin")
def export_all_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="all_payments.csv"'

    writer = csv.writer(response)
    writer.writerow(['Receipt No', 'Date', 'Time', 'Patient', 'Type', 'Doctor', 'Department', 'Amount'])

    data = get_all_payment_data(request)

    for row in data:
        writer.writerow([
            row["receipt_no"],
            row["date"].strftime("%d-%m-%Y"),
            row["time"],
            row["patient"],
            row["type"],
            row["doctor"],
            row["department"],
            row["amount"]
        ])

    return response


# ================= EXCEL =================
def export_all_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Payments"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D3B66")

    headers = ['Receipt No', 'Date', 'Time', 'Patient', 'Type', 'Doctor', 'Department', 'Amount (Rs)']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    data = get_all_payment_data(request)
    for row in data:
        ws.append([
            row["receipt_no"],
            row["date"].strftime("%d-%m-%Y"),
            row["time"],
            row["patient"],
            row["type"],
            row["doctor"],
            row["department"],
            row["amount"]
        ])

    col_widths = [14, 13, 10, 20, 12, 22, 35, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_payments.xlsx'
    wb.save(response)
    return response


# ================= PDF =================
def export_all_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.platypus import TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_payments.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(A4),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm
    )

    title_style = ParagraphStyle('t', fontSize=14, fontName='Helvetica-Bold', alignment=1, spaceAfter=6)
    sub_style   = ParagraphStyle('s', fontSize=9,  fontName='Helvetica',      alignment=1, spaceAfter=10)

    rows = get_all_payment_data(request)
    total = sum(r["amount"] for r in rows)
    from_date = request.GET.get("from_date", "")
    to_date   = request.GET.get("to_date", "")
    period = f"{from_date} to {to_date}" if from_date and to_date else "All Dates"

    table_data = [['Receipt No', 'Date', 'Time', 'Patient', 'Type', 'Doctor', 'Department', 'Amount (Rs)']]
    for r in rows:
        table_data.append([
            r["receipt_no"], r["date"].strftime("%d-%m-%Y"), r["time"],
            r["patient"], r["type"], r["doctor"], r["department"],
            f"{r['amount']:,.2f}"
        ])
    table_data.append(['', '', '', '', '', '', 'TOTAL', f"{total:,.2f}"])

    col_widths = [22*mm, 22*mm, 16*mm, 38*mm, 20*mm, 40*mm, 60*mm, 26*mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  colors.HexColor('#0D3B66')),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),  (-1,0),  8),
        ('ALIGN',         (0,0),  (-1,0),  'CENTER'),
        ('FONTSIZE',      (0,1),  (-1,-2), 7.5),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, colors.HexColor('#F0F4FF')]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor('#0D3B66')),
        ('TEXTCOLOR',     (0,-1), (-1,-1), colors.white),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN',         (-1,0), (-1,-1), 'RIGHT'),
        ('GRID',          (0,0),  (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
    ]))

    doc.build([
        Paragraph("Shradha Hospital & Multispeciality Centre", title_style),
        Paragraph(f"All Payment Receipts — {period}  |  Total: Rs.{total:,.2f}", sub_style),
        table
    ])
    return response

# ─────────────────────────────────────────────────────────────
# LIST
# ─────────────────────────────────────────────────────────────

@login_required
@role_required("admin")
def construction_expense_list(request):
    qs = ConstructionExpense.objects.select_related("vendor").order_by("-created_at")

    head      = request.GET.get("expense_head", "")
    area      = request.GET.get("area_location", "")
    status    = request.GET.get("approval_status", "")
    paid_by   = request.GET.get("paid_by", "")
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    q         = request.GET.get("q", "")

    if head:      qs = qs.filter(expense_head=head)
    if area:      qs = qs.filter(area_location=area)
    if status:    qs = qs.filter(approval_status=status)
    if paid_by:   qs = qs.filter(paid_by=paid_by)
    if date_from: qs = qs.filter(date__gte=date_from)
    if date_to:   qs = qs.filter(date__lte=date_to)
    if q:
        qs = qs.filter(
            Q(expense_id__icontains=q) | Q(description__icontains=q) |
            Q(vendor__name__icontains=q) | Q(bill_no__icontains=q)
        )

    totals = qs.aggregate(
        grand_total=Sum("total_amount"),
        total_gst=Sum("gst_amount"),
        count=Count("id"),
    )

    return render(request, "hms/construction/expense_list.html", {
        "expenses": qs,
        "totals": totals,
        "expense_head_choices": EXPENSE_HEAD_CHOICES,
        "area_choices": AREA_CHOICES,
        "approval_status_choices": APPROVAL_STATUS_CHOICES,
        "paid_by_choices": PAID_BY_CHOICES,
        "f_head": head, "f_area": area, "f_status": status,
        "f_paid_by": paid_by, "f_date_from": date_from,
        "f_date_to": date_to, "f_q": q,
    })


# ─────────────────────────────────────────────────────────────
# CREATE
# ─────────────────────────────────────────────────────────────

@login_required
@role_required("admin")
def construction_expense_create(request):
    vendors = Vendor.objects.all().order_by("name")

    if request.method == "POST":
        try:
            vendor_id   = request.POST.get("vendor_id", "").strip()
            vendor_name = request.POST.get("vendor_name", "").strip()
            vendor = None
            if vendor_id:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
            elif vendor_name:
                vendor, _ = Vendor.objects.get_or_create(
                    name=vendor_name,
                    defaults={"mobile": request.POST.get("vendor_mobile", "")}
                )

            expense = ConstructionExpense(
                date              = request.POST.get("date"),
                expense_head      = request.POST.get("expense_head"),
                subcategory       = request.POST.get("subcategory") or None,
                description       = request.POST.get("description"),
                area_location     = request.POST.get("area_location") or None,
                vendor            = vendor,
                vendor_mobile     = request.POST.get("vendor_mobile") or None,
                bill_no           = request.POST.get("bill_no") or None,
                qty               = request.POST.get("qty") or None,
                unit              = request.POST.get("unit") or None,
                rate              = request.POST.get("rate") or None,
                amount            = float(request.POST.get("amount") or 0),
                gst_percent       = float(request.POST.get("gst_percent") or 0),
                payment_mode      = request.POST.get("payment_mode") or None,
                paid_by           = request.POST.get("paid_by") or None,
                paid_from         = request.POST.get("paid_from") or None,
                approval_status   = request.POST.get("approval_status", "Pending"),
                approved_by       = request.POST.get("approved_by") or None,
                work_status       = request.POST.get("work_status") or None,
                material_received = request.POST.get("material_received") or None,
                invoice_type      = request.POST.get("invoice_type") or None,
                balance_due       = request.POST.get("balance_due") or None,
                due_date          = request.POST.get("due_date") or None,
                remarks           = request.POST.get("remarks") or None,
            )
            if request.FILES.get("bill_image"):
                expense.bill_image = request.FILES["bill_image"]
            if request.FILES.get("site_photo"):
                expense.site_photo = request.FILES["site_photo"]
            if request.FILES.get("quotation_file"):
                expense.quotation_file = request.FILES["quotation_file"]

            expense.save()
            messages.success(request, f"Expense {expense.expense_id} saved successfully.")
            return redirect("hms:construction_expense_list")

        except Exception as e:
            messages.error(request, f"Error saving expense: {e}")

    return render(request, "hms/construction/expense_form.html", {"vendors": vendors})


# ─────────────────────────────────────────────────────────────
# EDIT
# ─────────────────────────────────────────────────────────────

@login_required
@role_required("admin")
def construction_expense_edit(request, pk):
    expense = get_object_or_404(ConstructionExpense, pk=pk)
    vendors = Vendor.objects.all().order_by("name")

    if request.method == "POST":
        try:
            vendor_id   = request.POST.get("vendor_id", "").strip()
            vendor_name = request.POST.get("vendor_name", "").strip()
            vendor = None
            if vendor_id:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
            elif vendor_name:
                vendor, _ = Vendor.objects.get_or_create(
                    name=vendor_name,
                    defaults={"mobile": request.POST.get("vendor_mobile", "")}
                )

            expense.date              = request.POST.get("date")
            expense.expense_head      = request.POST.get("expense_head")
            expense.subcategory       = request.POST.get("subcategory") or None
            expense.description       = request.POST.get("description")
            expense.area_location     = request.POST.get("area_location") or None
            expense.vendor            = vendor
            expense.vendor_mobile     = request.POST.get("vendor_mobile") or None
            expense.bill_no           = request.POST.get("bill_no") or None
            expense.qty               = request.POST.get("qty") or None
            expense.unit              = request.POST.get("unit") or None
            expense.rate              = request.POST.get("rate") or None
            expense.amount            = float(request.POST.get("amount") or 0)
            expense.gst_percent       = float(request.POST.get("gst_percent") or 0)
            expense.payment_mode      = request.POST.get("payment_mode") or None
            expense.paid_by           = request.POST.get("paid_by") or None
            expense.paid_from         = request.POST.get("paid_from") or None
            expense.approval_status   = request.POST.get("approval_status", "Pending")
            expense.approved_by       = request.POST.get("approved_by") or None
            expense.work_status       = request.POST.get("work_status") or None
            expense.material_received = request.POST.get("material_received") or None
            expense.invoice_type      = request.POST.get("invoice_type") or None
            expense.balance_due       = request.POST.get("balance_due") or None
            expense.due_date          = request.POST.get("due_date") or None
            expense.remarks           = request.POST.get("remarks") or None

            if request.FILES.get("bill_image"):
                expense.bill_image = request.FILES["bill_image"]
            if request.FILES.get("site_photo"):
                expense.site_photo = request.FILES["site_photo"]
            if request.FILES.get("quotation_file"):
                expense.quotation_file = request.FILES["quotation_file"]

            expense.save()
            messages.success(request, f"Expense {expense.expense_id} updated.")
            return redirect("hms:construction_expense_list")

        except Exception as e:
            messages.error(request, f"Error updating: {e}")

    ctx = {"vendors": vendors}
    ctx.update({"expense": expense, "edit_mode": True})
    return render(request, "hms/construction/expense_form.html", ctx)


# ─────────────────────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────────────────────

@login_required
@role_required("admin")
def construction_expense_delete(request, pk):
    expense = get_object_or_404(ConstructionExpense, pk=pk)
    if request.method == "POST":
        eid = expense.expense_id
        expense.delete()
        messages.success(request, f"Expense {eid} deleted.")
    return redirect("hms:construction_expense_list")

@login_required
def construction_expense_receipt(request, pk):
    expense = get_object_or_404(ConstructionExpense, pk=pk)
    return render(request, "hms/construction/expense_receipt.html", {"expense": expense})

# ─────────────────────────────────────────────────────────────
# VENDOR AJAX
# ─────────────────────────────────────────────────────────────

@login_required
def vendor_search_ajax(request):
    q = request.GET.get("q", "")
    vendors = Vendor.objects.filter(name__icontains=q).values("id", "name", "mobile")[:10]
    return JsonResponse({"vendors": list(vendors)})

# ─────────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────────

def form_context(vendors):
    return {
        "vendors": vendors,
        "expense_head_choices": EXPENSE_HEAD_CHOICES,
        "area_choices": AREA_CHOICES,
        "payment_mode_choices": PAYMENT_MODE_CHOICES,
        "paid_by_choices": PAID_BY_CHOICES,
        "paid_from_choices": PAID_FROM_CHOICES,
        "approval_status_choices": APPROVAL_STATUS_CHOICES,
        "approved_by_choices": APPROVED_BY_CHOICES,
        "work_status_choices": WORK_STATUS_CHOICES,
        "material_received_choices": YES_NO_PARTIAL_CHOICES,
        "invoice_type_choices": INVOICE_TYPE_CHOICES,
        "today": date.today().strftime("%Y-%m-%d"),
    }

# ── USG: List all reports ────────────────────────────────────
@login_required
def usg_report_list(request):
    """List USG reports — with patient search filter."""
    q = request.GET.get("q", "").strip()
    reports = USGReport.objects.select_related("patient", "reporting_doctor")

    if q:
        reports = reports.filter(
            Q(patient__full_name__icontains=q) |
            Q(patient__uhid__icontains=q)      |
            Q(report_no__icontains=q)
        )

    return render(request, "hms/usg/usg_report_list.html", {
        "reports": reports[:100],
        "q":       q,
    })

# ── USG: Create new report ───────────────────────────────────────
@login_required
def usg_report_create(request, patient_id=None, bill_item_id=None):
    """Create a new USG report — optionally pre-linked to patient / bill item."""
    patient   = None
    bill_item = None

    if patient_id:
        patient = get_object_or_404(Patient, pk=patient_id)
    if bill_item_id:
        from .models import InvestigationBillItem
        bill_item = get_object_or_404(InvestigationBillItem, pk=bill_item_id)
        if not patient:
            patient = bill_item.bill.patient

    initial = {}
    if patient:
        initial["patient"] = patient
    if bill_item:
        initial["bill_item"] = bill_item

    form = USGReportForm(request.POST or None, initial=initial)

    if request.method == "POST" and form.is_valid():
        report            = form.save(commit=False)
        report.created_by = request.user
        report.save()
        messages.success(request, f"USG Report {report.report_no} saved successfully.")
        return redirect("hms:usg_report_print", pk=report.pk)

    return render(request, "hms/usg/usg_report_form.html", {
        "form":      form,
        "patient":   patient,
        "bill_item": bill_item,
        "title":     "New USG Report",
    })


# ── USG: Edit existing report ────────────────────────────────────
@login_required
def usg_report_edit(request, pk):
    report = get_object_or_404(USGReport, pk=pk)
    form   = USGReportForm(request.POST or None, instance=report)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Report {report.report_no} updated.")
        return redirect("hms:usg_report_print", pk=report.pk)

    return render(request, "hms/usg/usg_report_form.html", {
        "form":   form,
        "report": report,
        "title":  f"Edit {report.report_no}",
    })


# ── USG: Printable report ────────────────────────────────────────
@login_required
def usg_report_print(request, pk):
    """Render a print-ready USG report page."""
    report = get_object_or_404(
        USGReport.objects.select_related(
            "patient", "reporting_doctor", "referred_by", "consultation"
        ),
        pk=pk
    )
    return render(request, "hms/usg/usg_report_print.html", {
        "report": report,
    })


# ── USG: PDF download ────────────────────────────────────────────
@login_required
def usg_report_pdf(request, pk):
    """Generate PDF of USG report via xhtml2pdf."""
    report = get_object_or_404(
        USGReport.objects.select_related(
            "patient", "reporting_doctor", "referred_by"
        ),
        pk=pk
    )
    return render_to_pdf("hms/usg/usg_report_print.html", {"report": report})


# ── USG: Delete ──────────────────────────────────────────────────
@login_required 
def usg_report_delete(request, pk):
    report = get_object_or_404(USGReport, pk=pk)
    if request.method == "POST":
        report.delete()
        messages.success(request, "USG Report deleted.")
        return redirect("hms:usg_report_list")
    return render(request, "hms/usg/usg_report_confirm_delete.html", {"report": report})

from django.http import JsonResponse

@login_required
def patient_gender_api(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    return JsonResponse({"gender": patient.gender})

from django.http import JsonResponse

@login_required
def patient_get_api(request):
    patient_id = request.GET.get('id')
    try:
        p = Patient.objects.get(pk=patient_id)
        return JsonResponse({
            'id': p.pk,
            'full_name': p.full_name,
            'uhid': p.uhid,
        })
    except Patient.DoesNotExist:
        return JsonResponse({}, status=404)


@login_required
def patient_recent_consultation_api(request, patient_id):
    consultation = (
        Consultation.objects
        .filter(appointment__patient_id=patient_id)
        .select_related("diagnosis_icd")
        .prefetch_related("symptoms", "signs")
        .order_by("-created_at")
        .first()
    )
    if not consultation:
        return JsonResponse({"found": False})

    # Chief Complaints / Examination Findings are free-text fields that exist on
    # the model but the OPD consultation UI actually records this via the
    # symptoms/signs checklists (+ custom_symptoms/custom_signs) instead, so
    # pull from whichever of these is actually populated for this consultation.
    symptoms_lines = []
    if consultation.chief_complaints and consultation.chief_complaints.strip():
        symptoms_lines.append(consultation.chief_complaints.strip())

    symptom_names = list(consultation.symptoms.values_list("name", flat=True))
    if consultation.custom_symptoms and consultation.custom_symptoms.strip():
        symptom_names.append(consultation.custom_symptoms.strip())
    if symptom_names:
        symptoms_lines.append("Symptoms: " + ", ".join(symptom_names))

    sign_names = list(consultation.signs.values_list("name", flat=True))
    if consultation.custom_signs and consultation.custom_signs.strip():
        sign_names.append(consultation.custom_signs.strip())
    if sign_names:
        symptoms_lines.append("Signs: " + ", ".join(sign_names))

    if consultation.examination and consultation.examination.strip():
        symptoms_lines.append("Examination: " + consultation.examination.strip())

    diagnosis_text = (consultation.diagnosis_text or "").strip()
    if not diagnosis_text and consultation.diagnosis_icd:
        diagnosis_text = f"{consultation.diagnosis_icd.code} - {consultation.diagnosis_icd.description}"

    return JsonResponse({
        "found": True,
        "consultation_date": consultation.created_at.strftime("%d-%m-%Y"),
        "symptoms": "\n".join(symptoms_lines),
        "diagnosis": diagnosis_text,
    })


@login_required
def save_prescription_template(request):
    if request.method != 'POST': 
        return JsonResponse({'error': 'POST required'}, status=405)
    import json
    data      = json.loads(request.body)
    name      = data.get('name', '').strip()
    medicines = data.get('medicines', [])
    if not name or not medicines:
        return JsonResponse({'error': 'Name and medicines required'}, status=400)
    template, created = PrescriptionTemplate.objects.get_or_create(name=name)
    template.created_by = request.user
    template.save()
    template.items.all().delete()
    for i, m in enumerate(medicines):
        PrescriptionTemplateItem.objects.create(
            template=template, medicine=m.get('medicine',''),
            dose=m.get('dose',''), frequency=m.get('frequency',''),
            duration=m.get('duration',''), instructions=m.get('instructions',''),
            order=i
        )
    return JsonResponse({'success': True, 'id': template.pk, 'name': template.name, 'created': created})

@login_required
def list_prescription_templates(request):
    templates = PrescriptionTemplate.objects.all().order_by('-created_at')
    return JsonResponse({'templates': [{'id': t.pk, 'name': t.name} for t in templates]})

@login_required
def load_prescription_template(request, pk):
    try:
        t = PrescriptionTemplate.objects.get(pk=pk)
        items = [{'medicine': i.medicine, 'dose': i.dose, 'frequency': i.frequency,
                  'duration': i.duration, 'instructions': i.instructions} for i in t.items.all()]
        return JsonResponse({'id': t.pk, 'name': t.name, 'items': items})
    except PrescriptionTemplate.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

@login_required
def load_prescription_template(request, pk):
    try:
        tpl = PrescriptionTemplate.objects.get(pk=pk)
        items = tpl.items.all()
        return JsonResponse({
            'id': tpl.id,
            'name': tpl.name,
            'items': [
                {
                    'medicine': item.medicine,
                    'dose': item.dose,
                    'frequency': item.frequency,
                    'duration': item.duration,
                    'instructions': item.instructions,
                }
                for item in items
            ],
        })
    except PrescriptionTemplate.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
def delete_prescription_template(request, pk):
    try:
        PrescriptionTemplate.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except PrescriptionTemplate.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("doctor", "admin")
def add_symptom(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    name = data.get('name', '').strip()
    dept_id = data.get('department_id')

    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    sym, created = Symptom.objects.get_or_create(
        name=name,
        department_id=dept_id,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': sym.id, 'name': sym.name, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_symptom(request, pk):
    try:
        Symptom.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except Symptom.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("doctor", "admin")
def add_sign(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    name = data.get('name', '').strip()
    dept_id = data.get('department_id')

    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    sign, created = Sign.objects.get_or_create(
        name=name,
        department_id=dept_id,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': sign.id, 'name': sign.name, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_sign(request, pk):
    try:
        Sign.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except Sign.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("doctor", "admin")
def add_past_history(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    name = data.get('name', '').strip()

    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    item, created = PastHistory.objects.get_or_create(
        name=name,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': item.id, 'name': item.name, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_past_history(request, pk):
    try:
        PastHistory.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except PastHistory.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("doctor", "admin")
def add_surgical_history(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    name = data.get('name', '').strip()

    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    item, created = SurgicalHistory.objects.get_or_create(
        name=name,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': item.id, 'name': item.name, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_surgical_history(request, pk):
    try:
        SurgicalHistory.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except SurgicalHistory.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("reception", "admin", "doctor", "nursing")
def add_village(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    name = data.get('name', '').strip()

    if not name:
        return JsonResponse({'error': 'Name required'}, status=400)

    village, created = VillageMaster.objects.get_or_create(name=name)

    return JsonResponse({'id': village.id, 'name': village.name, 'created': created})


@login_required
@role_required("doctor", "admin")
def add_advice_option(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    text = data.get('text', '').strip()

    if not text:
        return JsonResponse({'error': 'Text required'}, status=400)

    item, created = AdviceOption.objects.get_or_create(
        text=text,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': item.id, 'text': item.text, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_advice_option(request, pk):
    try:
        AdviceOption.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except AdviceOption.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@role_required("doctor", "admin")
def add_diet_option(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data = json.loads(request.body)
    text = data.get('text', '').strip()

    if not text:
        return JsonResponse({'error': 'Text required'}, status=400)

    item, created = DietAdviceOption.objects.get_or_create(
        text=text,
        defaults={'is_active': True}
    )

    return JsonResponse({'id': item.id, 'text': item.text, 'created': created})


@login_required
@role_required("doctor", "admin")
def delete_diet_option(request, pk):
    try:
        DietAdviceOption.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except DietAdviceOption.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)



@login_required
def add_drug_quick(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        drug = DrugMaster.objects.create(
            name=data.get('name',''),
            generic_name=data.get('generic_name',''),
            strength=data.get('strength',''),
            category=data.get('category',''),
            default_dose=data.get('default_dose',''),
            default_frequency=data.get('default_frequency',''),
            default_duration=data.get('default_duration',''),
            default_instructions=data.get('default_instructions',''),
            is_active=True,
            sort_order=99,
        )
        return JsonResponse({'success': True, 'id': drug.id, 'name': drug.name})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def drug_defaults(request):
    name = request.GET.get('name', '').strip()
    if not name:
        return JsonResponse({'found': False})
    try:
        # Exact match first
        drug = DrugMaster.objects.filter(
            name__iexact=name, is_active=True
        ).first()
        if not drug:
            # Try first word match
            first_word = name.split()[0]
            drug = DrugMaster.objects.filter(
                name__istartswith=first_word, is_active=True
            ).first()
        if drug:
            return JsonResponse({
                'found': True,
                'dose': drug.default_dose or '',
                'frequency': drug.default_frequency or '',
                'duration': drug.default_duration or '',
                'instructions': drug.default_instructions or '',
                'generic_name': drug.generic_name or '',
                'atc_code': drug.atc_code or '',
            })
    except Exception:
        pass
    return JsonResponse({'found': False})
@login_required
def generate_diet(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not settings.AI_FEATURES_ENABLED:
        return JsonResponse({'error': 'AI features are not configured on this system.'})
    import json
    data = json.loads(request.body)
    diagnosis = data.get('diagnosis', '').strip()
    if not diagnosis:
        return JsonResponse({'error': 'No diagnosis'}, status=400)
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model='gpt-4o-mini',
            max_tokens=150,
            messages=[{
                'role': 'user',
                'content': f'Patient diagnosis: {diagnosis}.\nWrite Indian diet chart in exactly 6 lines:\nEarly Morning: ...\nBreakfast: ...\nLunch: ...\nEvening: ...\nDinner: ...\nAvoid: ...\nOnly diet chart, no extra text, 25 words max.'
            }]
        )
        diet = (response.choices[0].message.content or "").strip()
        return JsonResponse({'diet': diet})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def generate_lama_consent(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not settings.AI_FEATURES_ENABLED:
        return JsonResponse({'error': 'AI features are not configured on this system.'})
    import json
    data = json.loads(request.body)
    diagnosis = data.get('diagnosis', '').strip()
    plan = data.get('plan', '').strip()
    if not diagnosis or not plan:
        return JsonResponse({'error': 'Diagnosis and Plan/Advice are both required'}, status=400)
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model='gpt-4o-mini',
            max_tokens=900,
            response_format={
                "type": "json_schema",
                "json_schema": {
                    "name": "lama_consent",
                    "strict": True,
                    "schema": {
                        "type": "object",
                        "properties": {
                            "consent_en": {"type": "string"},
                            "consent_hi": {"type": "string"},
                        },
                        "required": ["consent_en", "consent_hi"],
                        "additionalProperties": False,
                    },
                },
            },
            messages=[
                {
                    'role': 'system',
                    'content': (
                        "You are a senior Indian surgeon drafting a formal LAMA "
                        "(Leave Against Medical Advice) consent note for a hospital "
                        "record. Always return ONLY valid JSON. No explanation."
                    ),
                },
                {
                    'role': 'user',
                    'content': f"""
Diagnosis: {diagnosis}
Recommended treatment / plan advised: {plan}

Write "consent_en": a formal 6-7 line consent/refusal paragraph in English,
first person plural ("We, the patient/attendant..."), covering:
- the diagnosis
- the treatment/admission that was recommended
- name 2-3 SPECIFIC, clinically accurate complications that could plausibly
  result from NOT receiving this exact treatment for this exact diagnosis
  (reason about the actual anatomy/pathology involved -- a fracture, a
  stone, an infection, etc. each have different realistic complications).
  Do not fall back to a generic, one-size-fits-all list of complications --
  they must be medically appropriate to THIS diagnosis specifically.
- a statement that the patient/attendant was informed of these risks in
  a language they understand, and voluntarily chose to decline/leave
  against medical advice, releasing the hospital and treating doctor
  from responsibility for adverse outcomes resulting from this decision.

Then write "consent_hi": a faithful Hindi translation of that exact
paragraph (Devanagari script), same meaning and structure.

Plain paragraph text only in each field, no headings, no markdown, no
bullet points.
""",
                },
            ],
        )
        raw = response.choices[0].message.content or "{}"
        parsed = json.loads(raw)
        return JsonResponse({
            'consent_en': parsed.get('consent_en', '').strip(),
            'consent_hi': parsed.get('consent_hi', '').strip(),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


from django.http import JsonResponse
from django.conf import settings
from django.contrib.auth.decorators import login_required
import json

@login_required
def generate_ai_medicines(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not settings.AI_FEATURES_ENABLED:
        return JsonResponse({'error': 'AI features are not configured on this system.'})

    try:
        data = json.loads(request.body)

        diagnosis = data.get('diagnosis', '').strip()
        drug_list = data.get('drug_list', [])

        if not diagnosis:
            return JsonResponse({'error': 'No diagnosis provided'}, status=400)

        # Convert drug list to string
        drug_str = ', '.join(drug_list) if drug_list else 'No drugs available'

        client = OpenAI(api_key=settings.OPENAI_API_KEY)

        # 🔥 Strong Prompt + JSON Enforcement
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=600,
            response_format={
                "type": "json_schema",
                "json_schema": {
                    "name": "ai_medicines",
                    "strict": True,
                    "schema": {
                        "type": "object",
                        "properties": {
                            "medicines": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "medicine": {"type": "string"},
                                        "dose": {"type": "string"},
                                        "frequency": {"type": "string"},
                                        "duration": {"type": "string"},
                                        "instructions": {"type": "string"},
                                    },
                                    "required": ["medicine", "dose", "frequency", "duration", "instructions"],
                                    "additionalProperties": False,
                                },
                            },
                        },
                        "required": ["medicines"],
                        "additionalProperties": False,
                    },
                },
            },
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a senior Indian physician. "
                        "Always return ONLY valid JSON. No explanation."
                    ),
                },
                {
                    "role": "user",
                    "content": f"""
Patient diagnosis: {diagnosis}

Available medicines in pharmacy:
{drug_str}

STRICT RULES:
- Prescribe only clinically correct medicines
- Prefer available medicines if appropriate
- If not available, write (not in stock)
- DO NOT add unnecessary medicines
- Maximum 3 to 5 medicines only

SURGICAL NOTE:
For appendicitis, hernia, cholelithiasis → give pre-operative medicines
(antibiotics, analgesics, antispasmodics, PPI)

DOSING RULES:
- Frequency must be fixed (e.g. Once daily (1-0-0), Twice daily (1-0-1))
- Duration must be fixed (e.g. 5 days, 7 days, 14 days)
- DO NOT use "as needed"

Return JSON format:
{{
  "medicines": [
    {{
      "medicine": "",
      "dose": "",
      "frequency": "",
      "duration": "",
      "instructions": ""
    }}
  ]
}}
"""
                }
            ]
        )

        # ✅ Direct JSON parsing (no regex)
        result_text = response.choices[0].message.content
        result = json.loads(result_text)

        # ✅ Validation layer
        validated_medicines = []
        required_keys = ["medicine", "dose", "frequency", "duration", "instructions"]

        for med in result.get("medicines", []):
            if all(key in med for key in required_keys):
                
                # Mark not in stock if not matched
                if drug_list and not any(med["medicine"].lower() in d.lower() for d in drug_list):
                    med["medicine"] += " (not in stock)"

                validated_medicines.append({
                    "medicine": med.get("medicine", ""),
                    "dose": med.get("dose", ""),
                    "frequency": med.get("frequency", ""),
                    "duration": med.get("duration", ""),
                    "instructions": med.get("instructions", "")
                })

        return JsonResponse({"medicines": validated_medicines})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON from AI'}, status=500)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
def partner_deposits(request):
    from .models import Partner
    partners = ['Pratap', 'Lumbaram', 'Poonaram']
    partner_labels = {
        'Pratap': 'Dr. Pratap',
        'Lumbaram': 'Mr. Lumbaram',
        'Poonaram': 'Mr. Poonaram',
    }

    if request.method == 'POST':
        partner_name = request.POST.get('partner')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        note = request.POST.get('note', '')
        if partner_name and amount and date:
            partner_obj, _ = Partner.objects.get_or_create(name=partner_name)
            PartnerDeposit.objects.create(
                partner=partner_obj,
                amount=amount,
                date=date,
                note=note,
                created_by=request.user,
            )
            messages.success(request, f"Deposit of ₹{amount} added for {partner_labels.get(partner_name, partner_name)}.")
        return redirect('hms:partner_deposits')

    deposit_sums = dict(
        PartnerDeposit.objects.values_list('partner__name')
        .annotate(total=Sum('amount'))
    )
    expense_sums = dict(
        ConstructionExpense.objects.values_list('paid_by')
        .annotate(total=Sum('total_amount'))
    )

    all_deposits = list(PartnerDeposit.objects.select_related('partner').order_by('-date'))
    all_expenses = list(ConstructionExpense.objects.order_by('-date'))

    summary = []
    for p in partners:
        total_deposited = deposit_sums.get(p) or Decimal('0')
        total_spent = expense_sums.get(p) or Decimal('0')
        summary.append({
            'key': p,
            'label': partner_labels[p],
            'deposited': total_deposited,
            'spent': total_spent,
            'balance': total_deposited - total_spent,
            'deposits': [d for d in all_deposits if d.partner.name == p],
            'expenses': [e for e in all_expenses if e.paid_by == p],
        })

    return render(request, 'hms/construction/partner_deposits.html', {
        'summary': summary,
    })
@login_required
def partner_deposit_edit(request, pk):
    from .models import PartnerDeposit
    deposit = get_object_or_404(PartnerDeposit, pk=pk)
    if request.method == 'POST':
        deposit.amount = request.POST.get('amount')
        deposit.date = request.POST.get('date')
        deposit.note = request.POST.get('note', '')
        deposit.save()
        messages.success(request, f"Deposit updated successfully.")
        return redirect('hms:partner_deposits')
    return redirect('hms:partner_deposits')


@login_required
def partner_deposit_delete(request, pk):
    from .models import PartnerDeposit
    deposit = get_object_or_404(PartnerDeposit, pk=pk)
    if request.method == 'POST':
        deposit.delete()
        messages.success(request, "Deposit deleted.")
    return redirect('hms:partner_deposits')